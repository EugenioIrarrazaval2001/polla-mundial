import os
import re
import json
import sys
import unicodedata
from openpyxl import Workbook, load_workbook
from datetime import datetime, date, time
from zoneinfo import ZoneInfo

# ============================================================
# CONFIG
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETA = BASE_DIR

OUTPUT_DIR = os.path.join(BASE_DIR, "site")
os.makedirs(OUTPUT_DIR, exist_ok=True)

CALENDARIO_PATH = os.path.join(BASE_DIR, "Calendario_Mundial.xlsx")
NOMBRES_PARTICIPANTES_PATH = os.path.join(BASE_DIR, "nombres_participantes.xlsx")

CELDA_INICIAL_RESULTADO = "B4"   # donde parte el partido 1
COL_MODO = "D"                   # en eliminatorias, el "cómo pasa" está en C4, C8, C12...
SALTO_FILAS = 4                  # cada partido está 4 filas más abajo

# Etapas:
# - tipo "GRUPOS": 1 celda por partido (resultado)
# - tipo "ELIM": 2 celdas por partido (pasa + modo)
ETAPAS = {
    "E01": {"tipo": "GRUPOS", "n_partidos": 25, "ppp": 1},
    "E02": {"tipo": "ELIM",   "n_partidos": 16, "ppp": 1},
    "E03": {"tipo": "ELIM",   "n_partidos": 8,  "ppp": 1},
    "E04": {"tipo": "ELIM",   "n_partidos": 4,  "ppp": 2},
    "E05": {"tipo": "ELIM",   "n_partidos": 2,  "ppp": 2},
    "E06": {"tipo": "ELIM",   "n_partidos": 1,  "ppp": 3},
}
ETIQUETAS_ETAPAS = {
    "E01": "Grupos",
    "E02": "16avos",
    "E03": "Octavos",
    "E04": "Cuartos",
    "E05": "Semis",
    "E06": "Final",
}
CALENDARIO_HOJAS = [
    ("E01", "E01_Grupos"),
    ("E02", "E02_16avos"),
    ("E03", "E03_Octavos"),
    ("E04", "E04_Cuartos"),
    ("E05", "E05_Semis"),
    ("E06", "E06_Final"),
]
CALENDARIO_HEADERS = [
    "numero_partido",
    "equipo_a",
    "equipo_b",
    "fecha_chile",
    "hora_chile",
    "datetime_chile_iso",
    "sede",
    "notas",
]
CALENDARIO_E01_INICIAL = [
    (1, "México", "Sudáfrica", "2026-06-11", "15:00", "2026-06-11T15:00:00-04:00"),
    (2, "Estados Unidos", "Paraguay", "2026-06-12", "21:00", "2026-06-12T21:00:00-04:00"),
    (3, "Brasil", "Marruecos", "2026-06-13", "18:00", "2026-06-13T18:00:00-04:00"),
    (4, "Australia", "Turquía", "2026-06-14", "00:00", "2026-06-14T00:00:00-04:00"),
    (5, "Holanda", "Japón", "2026-06-14", "16:00", "2026-06-14T16:00:00-04:00"),
    (6, "Bélgica", "Egipto", "2026-06-15", "15:00", "2026-06-15T15:00:00-04:00"),
    (7, "Francia", "Senegal", "2026-06-16", "15:00", "2026-06-16T15:00:00-04:00"),
    (8, "Argentina", "Argelia", "2026-06-16", "21:00", "2026-06-16T21:00:00-04:00"),
    (9, "Inglaterra", "Croacia", "2026-06-17", "16:00", "2026-06-17T16:00:00-04:00"),
    (10, "México", "Corea", "2026-06-18", "21:00", "2026-06-18T21:00:00-04:00"),
    (11, "Estados Unidos", "Australia", "2026-06-19", "15:00", "2026-06-19T15:00:00-04:00"),
    (12, "Holanda", "Suecia", "2026-06-20", "13:00", "2026-06-20T13:00:00-04:00"),
    (13, "Bélgica", "Irán", "2026-06-21", "15:00", "2026-06-21T15:00:00-04:00"),
    (14, "Argentina", "Austria", "2026-06-22", "13:00", "2026-06-22T13:00:00-04:00"),
    (15, "Noruega", "Senegal", "2026-06-22", "20:00", "2026-06-22T20:00:00-04:00"),
    (16, "Suiza", "Canadá", "2026-06-24", "15:00", "2026-06-24T15:00:00-04:00"),
    (17, "Ecuador", "Alemania", "2026-06-25", "16:00", "2026-06-25T16:00:00-04:00"),
    (18, "Japón", "Suecia", "2026-06-25", "19:00", "2026-06-25T19:00:00-04:00"),
    (19, "Turquía", "Estados Unidos", "2026-06-25", "22:00", "2026-06-25T22:00:00-04:00"),
    (20, "Paraguay", "Australia", "2026-06-25", "22:00", "2026-06-25T22:00:00-04:00"),
    (21, "Uruguay", "España", "2026-06-26", "20:00", "2026-06-26T20:00:00-04:00"),
    (22, "Noruega", "Francia", "2026-06-26", "15:00", "2026-06-26T15:00:00-04:00"),
    (23, "Egipto", "Irán", "2026-06-26", "23:00", "2026-06-26T23:00:00-04:00"),
    (24, "Argelia", "Austria", "2026-06-27", "22:00", "2026-06-27T22:00:00-04:00"),
    (25, "Colombia", "Portugal", "2026-06-27", "19:30", "2026-06-27T19:30:00-04:00"),
]
DIAS_DISPLAY = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
MESES_DISPLAY = {
    1: "ene",
    2: "feb",
    3: "mar",
    4: "abr",
    5: "may",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "sep",
    10: "oct",
    11: "nov",
    12: "dic",
}

def clave_orden_etapa(etapa):
    m = re.match(r"^E(\d{2})$", str(etapa).upper())
    return int(m.group(1)) if m else 9999


def etiqueta_etapa_larga(etapa):
    nombre = ETIQUETAS_ETAPAS.get(etapa, etapa)
    if str(etapa).upper() == "E01":
        nombre = "Fase de grupos"
    return nombre


def texto_celda(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def hoja_calendario_vacia(ws):
    for row in ws.iter_rows():
        for cell in row:
            if texto_celda(cell.value):
                return False
    return True


def calendario_tiene_filas_datos(ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(texto_celda(valor) for valor in row):
            return True
    return False


def asegurar_headers_calendario(ws):
    if hoja_calendario_vacia(ws):
        ws.append(CALENDARIO_HEADERS)
        return True
    return False


def agregar_e01_inicial_si_corresponde(ws):
    if calendario_tiene_filas_datos(ws):
        return False

    for numero, equipo_a, equipo_b, fecha, hora, iso in CALENDARIO_E01_INICIAL:
        ws.append([numero, equipo_a, equipo_b, fecha, hora, iso, "", ""])
    return True


def normalizar_numero_partido(valor):
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        numero = int(valor)
        return numero if numero == valor or float(numero) == float(valor) else None

    texto = texto_celda(valor)
    if not texto:
        return None
    m = re.search(r"\d+", texto)
    return int(m.group(0)) if m else None


def normalizar_fecha_calendario(valor):
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")

    texto = texto_celda(valor)
    if not texto:
        return ""

    formatos = ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y")
    for formato in formatos:
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return texto


def normalizar_hora_calendario(valor):
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        if 0 <= valor < 1:
            minutos = int(round(valor * 24 * 60))
            horas = (minutos // 60) % 24
            mins = minutos % 60
            return f"{horas:02d}:{mins:02d}"
        if 0 <= valor < 24:
            horas = int(valor)
            mins = int(round((valor - horas) * 60))
            return f"{horas:02d}:{mins:02d}"

    texto = texto_celda(valor)
    if not texto:
        return ""
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", texto)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return texto


def parse_datetime_iso_calendario(valor):
    if isinstance(valor, datetime):
        dt = valor
    else:
        texto = texto_celda(valor)
        if not texto:
            return None
        try:
            dt = datetime.fromisoformat(texto.replace("Z", "+00:00"))
        except ValueError:
            return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("America/Santiago"))
    return dt


def normalizar_iso_calendario(valor):
    dt = parse_datetime_iso_calendario(valor)
    if dt:
        return dt.isoformat(timespec="seconds")
    return texto_celda(valor)


def construir_iso_calendario(fecha_chile, hora_chile):
    if not fecha_chile or not hora_chile:
        return ""
    try:
        dt = datetime.strptime(f"{fecha_chile} {hora_chile}", "%Y-%m-%d %H:%M")
    except ValueError:
        return ""
    dt = dt.replace(tzinfo=ZoneInfo("America/Santiago"))
    return dt.isoformat(timespec="seconds")


def datetime_calendario_para_display(fecha_chile, hora_chile, datetime_chile_iso):
    dt = parse_datetime_iso_calendario(datetime_chile_iso)
    if dt:
        return dt
    if fecha_chile and hora_chile:
        try:
            return datetime.strptime(f"{fecha_chile} {hora_chile}", "%Y-%m-%d %H:%M")
        except ValueError:
            return None
    return None


def formatear_fecha_hora_display(fecha_chile, hora_chile, datetime_chile_iso):
    dt = datetime_calendario_para_display(fecha_chile, hora_chile, datetime_chile_iso)
    if not dt:
        return "Horario por confirmar"

    hora_display = hora_chile or dt.strftime("%H:%M")
    mes_display = MESES_DISPLAY.get(dt.month, "")
    return f"{DIAS_DISPLAY[dt.weekday()]} {dt.day} {mes_display} · {hora_display}"


def preparar_y_cargar_calendario():
    if os.path.exists(CALENDARIO_PATH):
        wb = load_workbook(CALENDARIO_PATH)
        cambiado = False
    else:
        wb = Workbook()
        wb.active.title = CALENDARIO_HOJAS[0][1]
        cambiado = True

    for idx, (etapa, nombre_hoja) in enumerate(CALENDARIO_HOJAS):
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            ws = wb.create_sheet(title=nombre_hoja, index=idx)
            cambiado = True

        if asegurar_headers_calendario(ws):
            cambiado = True
        if etapa == "E01" and agregar_e01_inicial_si_corresponde(ws):
            cambiado = True

    if cambiado:
        wb.save(CALENDARIO_PATH)

    calendario = {etapa: {} for etapa, _ in CALENDARIO_HOJAS}
    hoja_por_etapa = dict(CALENDARIO_HOJAS)

    for etapa, nombre_hoja in hoja_por_etapa.items():
        if nombre_hoja not in wb.sheetnames:
            continue

        ws = wb[nombre_hoja]
        headers = {
            texto_celda(cell.value).lower(): idx
            for idx, cell in enumerate(ws[1])
            if texto_celda(cell.value)
        }
        if "numero_partido" not in headers:
            continue

        def valor_columna(row, nombre):
            idx = headers.get(nombre)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(texto_celda(valor) for valor in row):
                continue

            numero = normalizar_numero_partido(valor_columna(row, "numero_partido"))
            if numero is None:
                continue

            fecha_chile = normalizar_fecha_calendario(valor_columna(row, "fecha_chile"))
            hora_chile = normalizar_hora_calendario(valor_columna(row, "hora_chile"))
            datetime_chile_iso = normalizar_iso_calendario(valor_columna(row, "datetime_chile_iso"))
            if not datetime_chile_iso and fecha_chile and hora_chile:
                datetime_chile_iso = construir_iso_calendario(fecha_chile, hora_chile)

            calendario[etapa][numero] = {
                "equipo_a": texto_celda(valor_columna(row, "equipo_a")),
                "equipo_b": texto_celda(valor_columna(row, "equipo_b")),
                "fecha_chile": fecha_chile,
                "hora_chile": hora_chile,
                "datetime_chile_iso": datetime_chile_iso,
                "fecha_hora_display": formatear_fecha_hora_display(
                    fecha_chile,
                    hora_chile,
                    datetime_chile_iso,
                ),
                "sort_key": datetime_chile_iso,
                "sede": texto_celda(valor_columna(row, "sede")),
                "notas": texto_celda(valor_columna(row, "notas")),
            }

    return calendario

# ============================================================
# BONUS CAMPEÓN
# - Cada participante escribe en B4 de su archivo E01 su campeón.
# - El campeón real OFICIAL se lee desde B4 del archivo de pauta E01.
# - CAMPEON_REAL_MANUAL queda solo como fallback opcional.
# ============================================================
BONUS_PTS = 5
NOMBRE_COLUMNA_BONUS = "Bono Campeón"
CAMPEON_REAL_MANUAL = None
PREMIOS_POR_POSICION = {
    1: 60,
    2: 30,
    3: 10,
}

if sum(PREMIOS_POR_POSICION.values()) != 100:
    raise ValueError("La distribución de premios debe sumar 100%.")


# ============================================================
# DETECCIÓN DE ETAPA / PARTICIPANTE
# ============================================================
def extraer_etapa_desde_texto(texto):
    limpio = str(texto).strip()

    # E01 / E1 / e01
    m = re.search(r"(?i)\bE\s*0*(\d{1,2})\b", limpio)
    if m:
        etapa = f"E{int(m.group(1)):02d}"
        return etapa if etapa in ETAPAS else None

    # etapa 01 / etapa_1 / Etapa-02
    m = re.search(r"(?i)\bETAPA\D*0*(\d{1,2})\b", limpio)
    if m:
        etapa = f"E{int(m.group(1)):02d}"
        return etapa if etapa in ETAPAS else None

    # Solo número: 01 / 1
    m = re.fullmatch(r"0*(\d{1,2})", limpio)
    if m:
        etapa = f"E{int(m.group(1)):02d}"
        return etapa if etapa in ETAPAS else None

    return None


def extraer_etapa_desde_nombre(fn):
    base = os.path.splitext(os.path.basename(fn))[0]
    # En pauta permitimos prefijos tipo E01Pauta.xlsx (sin separador).
    m = re.match(r"(?i)^E\s*0*(\d{1,2})", base)
    if m:
        etapa = f"E{int(m.group(1)):02d}"
        return etapa if etapa in ETAPAS else None
    return extraer_etapa_desde_texto(base)


def extraer_etapa_desde_ruta_participante(ruta_excel, carpeta_participantes):
    carpeta_archivo = os.path.dirname(os.path.abspath(ruta_excel))
    rel = os.path.relpath(carpeta_archivo, os.path.abspath(carpeta_participantes))
    if rel in (".", ""):
        return None

    partes = rel.split(os.sep)
    for parte in reversed(partes):
        etapa = extraer_etapa_desde_texto(parte)
        if etapa:
            return etapa
    return None


def nombre_participante_desde_archivo(ruta_excel):
    base = os.path.splitext(os.path.basename(ruta_excel))[0]
    # Acepta nombres simples con guiones/guiones bajos.
    nombre = re.sub(r"[_-]+", " ", base)
    nombre = re.sub(r"\s+", " ", nombre).strip()
    return nombre


def id_participante(nombre):
    # ID estable derivada del nombre de archivo (sin convención antigua).
    return normalizar_texto(nombre)


def resolver_carpeta_por_nombre(carpeta_base, nombre_objetivo):
    objetivo = str(nombre_objetivo).strip().lower()
    for nombre in os.listdir(carpeta_base):
        ruta = os.path.join(carpeta_base, nombre)
        if os.path.isdir(ruta) and nombre.lower() == objetivo:
            return ruta
    raise FileNotFoundError(
        f"No encontré la carpeta '{nombre_objetivo}' dentro de: {carpeta_base}"
    )


def resolver_carpeta_pauta(carpeta_base):
    # Busca "pauta" sin depender de mayúsculas/minúsculas.
    return resolver_carpeta_por_nombre(carpeta_base, "Pauta")


def cargar_pautas_desde_excel(carpeta_pauta):
    archivos_excel = sorted(
        fn for fn in os.listdir(carpeta_pauta)
        if fn.lower().endswith(".xlsx") and not fn.startswith("~$")
    )
    if not archivos_excel:
        raise FileNotFoundError(
            f"La carpeta de pauta está vacía o no tiene .xlsx: {carpeta_pauta}"
        )

    pautas = {}
    fuentes = {}
    ignorados = []
    enfrentamientos = {}
    enfrentamientos_detalle = {}
    campeon_real_pauta = None

    for fn in archivos_excel:
        etapa = extraer_etapa_desde_nombre(fn)
        if etapa is None:
            ignorados.append(fn)
            continue
        if etapa not in ETAPAS:
            ignorados.append(fn)
            continue
        if etapa in pautas:
            raise ValueError(
                f"Hay más de una pauta para {etapa}: '{fuentes[etapa]}' y '{fn}'. "
                "Deja solo un archivo por etapa en la carpeta pauta."
            )

        ruta_excel = os.path.join(carpeta_pauta, fn)
        cfg = ETAPAS[etapa]
        wb = load_workbook(ruta_excel, data_only=True)
        ws = wb.active

        if etapa == "E01":
            campeon_real_pauta = ws["B4"].value

        if cfg["tipo"] == "GRUPOS":
            pauta_etapa = leer_celdas_resultado(ws, cfg["n_partidos"])
        elif cfg["tipo"] == "ELIM":
            pauta_etapa = leer_celdas_eliminatoria(ws, cfg["n_partidos"])
        else:
            raise ValueError(f"Tipo de etapa desconocido: {cfg['tipo']}")
        enfrentamientos_detalle_etapa = leer_enfrentamientos_etapa_detalle(ws, cfg)
        enfrentamientos_etapa = [p["nombre"] for p in enfrentamientos_detalle_etapa]

        if len(pauta_etapa) != cfg["n_partidos"]:
            raise ValueError(
                f"La pauta {fn} ({etapa}) no tiene {cfg['n_partidos']} partidos."
            )

        pautas[etapa] = pauta_etapa
        fuentes[etapa] = fn
        enfrentamientos[etapa] = enfrentamientos_etapa
        enfrentamientos_detalle[etapa] = enfrentamientos_detalle_etapa

    if not pautas:
        raise ValueError(
            f"No pude cargar pautas válidas desde: {carpeta_pauta}. "
            "Asegúrate de usar archivos cuyo nombre comience por etapa (ej: E01...)."
        )

    faltantes = [e for e in sorted(ETAPAS.keys()) if e not in pautas]
    return pautas, fuentes, faltantes, ignorados, enfrentamientos, campeon_real_pauta, enfrentamientos_detalle


def cargar_archivos_pronostico(carpeta_participantes):
    registros = []
    avisos = []
    etapas_con_carpeta = set()
    conteo_por_etapa = {}

    for raiz, dirs, files in os.walk(carpeta_participantes):
        dirs.sort()
        files = sorted(files)

        etapa_en_raiz = extraer_etapa_desde_texto(os.path.basename(raiz))
        if etapa_en_raiz in ETAPAS:
            etapas_con_carpeta.add(etapa_en_raiz)

        for fn in files:
            if fn.startswith("~$") or not fn.lower().endswith(".xlsx"):
                continue

            ruta = os.path.join(raiz, fn)
            etapa = extraer_etapa_desde_ruta_participante(ruta, carpeta_participantes)
            if etapa is None:
                avisos.append(
                    f"Archivo ignorado (no pude inferir etapa por carpeta): {ruta}"
                )
                continue
            if etapa not in ETAPAS:
                avisos.append(
                    f"Archivo ignorado (etapa fuera de configuración): {ruta}"
                )
                continue

            nombre = nombre_participante_desde_archivo(ruta)
            if not nombre:
                avisos.append(
                    f"Archivo ignorado (nombre de participante vacío): {ruta}"
                )
                continue

            registros.append({
                "ruta": ruta,
                "archivo": fn,
                "etapa": etapa,
                "nombre": nombre,
                "pid": id_participante(nombre),
            })
            conteo_por_etapa[etapa] = conteo_por_etapa.get(etapa, 0) + 1

    etapas_vacias = [
        e for e in sorted(etapas_con_carpeta, key=clave_orden_etapa)
        if conteo_por_etapa.get(e, 0) == 0
    ]

    registros.sort(key=lambda r: (clave_orden_etapa(r["etapa"]), r["nombre"], r["archivo"]))
    return registros, avisos, etapas_vacias


# ============================================================
# LECTURA CELDAS
# ============================================================
def leer_celdas_resultado(ws, n_partidos, celda_inicial=CELDA_INICIAL_RESULTADO, salto_filas=SALTO_FILAS):
    col = celda_inicial[0]
    fila = int(celda_inicial[1:]) + 4  # (mantengo EXACTO tu offset)
    out = []
    for _ in range(n_partidos):
        out.append(ws[f"{col}{fila}"].value)
        fila += salto_filas
    return out


def leer_celdas_eliminatoria(ws, n_partidos, celda_inicial=CELDA_INICIAL_RESULTADO, col_modo=COL_MODO, salto_filas=SALTO_FILAS):
    col_pasa = celda_inicial[0]
    fila = int(celda_inicial[1:])
    out = []
    for _ in range(n_partidos):
        pasa = ws[f"{col_pasa}{fila}"].value
        modo = ws[f"{col_modo}{fila}"].value
        out.append((pasa, modo))
        fila += salto_filas
    return out


def leer_campeon_predicho_desde_e01(ruta_excel):
    """
    En E01, el campeón pronosticado está en B4.
    """
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active
    return ws["B4"].value


# ============================================================
# PUNTAJE
# ============================================================
def normalizar_texto(x):
    if x is None:
        return ""
    return " ".join(str(x).strip().split()).upper()


def texto_ganador_sin_prefijo(x):
    txt = "" if x is None else str(x).strip()
    if not txt:
        return ""
    return re.sub(r"(?i)^\s*(?:pasa|gana|clasifica|avanza)\b[\s:.-]*", "", txt, count=1).strip()


def texto_pasa_eliminatoria(x):
    return texto_ganador_sin_prefijo(x)


def normalizar_pasa_eliminatoria(x):
    return normalizar_texto(texto_pasa_eliminatoria(x))


def puntaje_grupos(apuestas, pauta):
    pts = 0
    for a, r in zip(apuestas, pauta):
        if normalizar_texto(a) == normalizar_texto(r):
            pts += 1
    return pts


def puntaje_eliminatoria(apuestas, pauta, ppp):
    """
    Regla:
    - Si falla "pasa": 0
    - Si acierta "pasa": suma ppp
    - Si además acierta "modo": suma +1 (tal como lo tienes ahora)
    """
    pts = 0

    for (pasa, modo), (pasa_real, modo_real) in zip(apuestas, pauta):
        if normalizar_pasa_eliminatoria(pasa) != normalizar_pasa_eliminatoria(pasa_real):
            continue
        pts += ppp
        if normalizar_texto(modo) == normalizar_texto(modo_real):
            pts += 1

    return pts


def valor_visible(x):
    if x is None:
        return "-"
    txt = str(x).strip()
    return txt if txt else "-"


def valor_payload(x):
    if x is None:
        return ""
    return str(x).strip()


def pauta_partido_tiene_resultado(partido):
    if isinstance(partido, (list, tuple)):
        return any(normalizar_texto(valor) != "" for valor in partido)
    return normalizar_texto(partido) != ""


def etapa_comenzada(pautas_por_etapa, etapa):
    pauta = pautas_por_etapa.get(etapa)
    if not pauta:
        return False
    return any(pauta_partido_tiene_resultado(partido) for partido in pauta)


def pauta_partido_finalizado(etapa, partido):
    cfg = ETAPAS[etapa]
    if cfg["tipo"] == "GRUPOS":
        return normalizar_texto(partido) != ""

    if cfg["tipo"] == "ELIM":
        if isinstance(partido, (list, tuple)):
            pasa_real = partido[0] if len(partido) > 0 else None
            modo_real = partido[1] if len(partido) > 1 else None
        else:
            pasa_real = partido
            modo_real = None
        return normalizar_pasa_eliminatoria(pasa_real) != "" and normalizar_texto(modo_real) != ""

    return False


def etapa_finalizada(pautas_por_etapa, etapa):
    pauta = pautas_por_etapa.get(etapa)
    if not pauta:
        return False

    cfg = ETAPAS[etapa]
    if len(pauta) < cfg["n_partidos"]:
        return False

    return all(
        pauta_partido_finalizado(etapa, partido)
        for partido in pauta[:cfg["n_partidos"]]
    )


def normalizar_comparacion(x):
    txt = normalizar_texto(x)
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    txt = re.sub(r"[^A-Z0-9]+", " ", txt)
    return " ".join(txt.split())


def cargar_mapa_familias(ruta_excel=NOMBRES_PARTICIPANTES_PATH):
    """Carga participantes y familias preservando los nombres visibles del Excel."""
    if not os.path.exists(ruta_excel):
        raise FileNotFoundError(f"no existe el archivo: {ruta_excel}")

    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    try:
        ws = wb.active
        mapa_participantes = {}
        familias = {}
        avisos = []
        encabezados_nombre = {
            "NOMBRE", "NOMBRES", "PARTICIPANTE", "PARTICIPANTES",
            "NOMBRE PARTICIPANTE", "NOMBRE DEL PARTICIPANTE",
        }
        encabezados_familia = {
            "FAMILIA", "FAMILIAS", "GRUPO", "NOMBRE FAMILIA",
            "NOMBRE DE FAMILIA", "NOMBRE DE LA FAMILIA",
        }

        for numero_fila, (nombre_raw, familia_raw) in enumerate(
            ws.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True),
            start=1,
        ):
            nombre = str(nombre_raw).strip() if nombre_raw is not None else ""
            familia = str(familia_raw).strip() if familia_raw is not None else ""
            if not nombre and not familia:
                continue

            nombre_norm = normalizar_comparacion(nombre)
            familia_norm = normalizar_comparacion(familia)
            if nombre_norm in encabezados_nombre and familia_norm in encabezados_familia:
                continue

            if not nombre or not familia:
                avisos.append(
                    f"fila {numero_fila} ignorada: debe tener nombre y familia en columnas A y B."
                )
                continue
            if not nombre_norm or not familia_norm:
                avisos.append(
                    f"fila {numero_fila} ignorada: nombre o familia sin caracteres comparables."
                )
                continue

            asignacion_anterior = mapa_participantes.get(nombre_norm)
            if asignacion_anterior:
                if asignacion_anterior["familia_norm"] != familia_norm:
                    raise ValueError(
                        "asignación familiar contradictoria para "
                        f"'{nombre}': '{asignacion_anterior['familia']}' y '{familia}' "
                        f"(fila {numero_fila})."
                    )
                continue

            familia_info = familias.setdefault(familia_norm, {
                "nombre": familia,
                "integrantes_excel": [],
            })
            familia_info["integrantes_excel"].append(nombre)
            mapa_participantes[nombre_norm] = {
                "nombre_excel": nombre,
                "familia": familia_info["nombre"],
                "familia_norm": familia_norm,
            }

        return {
            "participantes": mapa_participantes,
            "familias": familias,
            "avisos": avisos,
        }
    finally:
        wb.close()


def es_empate_pauta(x):
    return normalizar_comparacion(x) in {
        "EMPATE",
        "EMPATADO",
        "E",
        "DRAW",
        "TIE",
        "X",
        "IGUALDAD",
    }


def etiqueta_modo_eliminatoria(modo):
    txt = valor_payload(modo)
    if not txt:
        return ""

    norm = normalizar_comparacion(txt)
    if norm in {"90", "90 MIN", "90 MINUTOS", "TIEMPO REGULAR", "REGULAR", "EN LOS 90", "EN 90", "LOS 90"}:
        return "90'"
    if norm in {"ALARGUE", "ALARGUE EXTRA", "TIEMPO EXTRA", "EN TIEMPO EXTRA", "SUPLEMENTARIO", "PRORROGA"}:
        return "Alargue"
    if norm in {"PENALES", "EN PENALES", "PENAL", "PENALTIS", "PENALTIES", "PK", "PEN"}:
        return "Penales"
    return txt


def formatear_prediccion_elim(pasa, modo):
    vp = valor_visible(texto_pasa_eliminatoria(pasa))
    vm = valor_visible(modo)
    if vp == "-" and vm == "-":
        return "-"
    if vm == "-":
        return vp
    return f"{vp} | {vm}"


def formatear_enfrentamiento(equipo_a, equipo_b, numero_partido):
    a = valor_visible(equipo_a)
    b = valor_visible(equipo_b)
    if a != "-" and b != "-":
        return f"{a} vs {b}"
    if a != "-":
        return a
    if b != "-":
        return b
    return f"Partido {numero_partido}"


def leer_enfrentamientos_etapa_detalle(ws, cfg, celda_inicial=CELDA_INICIAL_RESULTADO, salto_filas=SALTO_FILAS):
    col_equipo_a = celda_inicial[0]
    col_equipo_b = "D"
    fila_base = int(celda_inicial[1:])
    if cfg["tipo"] == "GRUPOS":
        fila_prediccion = fila_base + 4  # mismo offset usado en leer_celdas_resultado
    elif cfg["tipo"] == "ELIM":
        fila_prediccion = fila_base
    else:
        raise ValueError(f"Tipo de etapa desconocido: {cfg['tipo']}")

    out = []
    for i in range(cfg["n_partidos"]):
        numero_partido = i + 1
        fila_enfrentamiento = (fila_prediccion - 1) + (i * salto_filas)
        equipo_a = ws[f"{col_equipo_a}{fila_enfrentamiento}"].value
        equipo_b = ws[f"{col_equipo_b}{fila_enfrentamiento}"].value
        out.append({
            "numero": numero_partido,
            "nombre": formatear_enfrentamiento(equipo_a, equipo_b, numero_partido),
            "equipo_a": valor_payload(equipo_a),
            "equipo_b": valor_payload(equipo_b),
        })
    return out


def leer_enfrentamientos_etapa(ws, cfg, celda_inicial=CELDA_INICIAL_RESULTADO, salto_filas=SALTO_FILAS):
    return [
        partido["nombre"]
        for partido in leer_enfrentamientos_etapa_detalle(ws, cfg, celda_inicial, salto_filas)
    ]


def interpretar_resultado_grupos(pauta, equipo_a, equipo_b):
    resultado = valor_payload(pauta)
    if not resultado:
        return {
            "resultado": "Pendiente",
            "ganador": "",
            "winner_side": "",
            "modo": "",
            "estado": "pendiente",
            "outcome": "pending",
        }

    ganador_sin_prefijo = texto_ganador_sin_prefijo(resultado)
    norm_resultado = normalizar_comparacion(ganador_sin_prefijo)
    norm_a = normalizar_comparacion(equipo_a)
    norm_b = normalizar_comparacion(equipo_b)

    if norm_a and norm_resultado == norm_a:
        return {
            "resultado": f"Gana {valor_payload(equipo_a)}",
            "ganador": valor_payload(equipo_a),
            "winner_side": "A",
            "modo": "",
            "estado": "jugado",
            "outcome": "winner",
        }
    if norm_b and norm_resultado == norm_b:
        return {
            "resultado": f"Gana {valor_payload(equipo_b)}",
            "ganador": valor_payload(equipo_b),
            "winner_side": "B",
            "modo": "",
            "estado": "jugado",
            "outcome": "winner",
        }
    if es_empate_pauta(resultado):
        return {
            "resultado": "Empate",
            "ganador": "",
            "winner_side": "",
            "modo": "",
            "estado": "jugado",
            "outcome": "draw",
        }

    return {
        "resultado": resultado,
        "ganador": "",
        "winner_side": "",
        "modo": "",
        "estado": "jugado",
        "outcome": "text",
    }


def interpretar_resultado_eliminatoria(pauta, equipo_a, equipo_b):
    if isinstance(pauta, (list, tuple)):
        pasa = pauta[0] if len(pauta) > 0 else None
        modo = pauta[1] if len(pauta) > 1 else None
    else:
        pasa = pauta
        modo = None

    ganador = texto_pasa_eliminatoria(pasa)
    modo_txt = etiqueta_modo_eliminatoria(modo)
    if not ganador:
        return {
            "resultado": "Pendiente",
            "ganador": "",
            "winner_side": "",
            "modo": modo_txt,
            "estado": "pendiente",
            "outcome": "pending",
        }

    norm_ganador = normalizar_comparacion(ganador)
    norm_a = normalizar_comparacion(equipo_a)
    norm_b = normalizar_comparacion(equipo_b)
    winner_side = ""
    if norm_a and norm_ganador == norm_a:
        winner_side = "A"
    elif norm_b and norm_ganador == norm_b:
        winner_side = "B"

    resultado = f"Pasa {ganador}"
    if modo_txt:
        resultado = f"{resultado} - {modo_txt}"

    return {
        "resultado": resultado,
        "ganador": ganador,
        "winner_side": winner_side,
        "modo": modo_txt,
        "estado": "jugado",
        "outcome": "winner",
    }


def construir_resultados_payload(etapas_ordenadas, pautas_por_etapa, enfrentamientos_detalle_por_etapa, calendario_por_etapa=None):
    stages = []
    matches = {}
    calendario_por_etapa = calendario_por_etapa or {}

    for etapa in etapas_ordenadas:
        cfg = ETAPAS[etapa]
        stages.append({
            "id": etapa,
            "label": etiqueta_etapa_larga(etapa),
            "type": cfg["tipo"],
        })

        pauta_etapa = pautas_por_etapa.get(etapa)
        enfrentamientos = enfrentamientos_detalle_por_etapa.get(etapa, [])
        calendario_etapa = calendario_por_etapa.get(etapa, {})
        partidos = []

        if pauta_etapa is None:
            matches[etapa] = partidos
            continue

        for i in range(cfg["n_partidos"]):
            numero_partido = i + 1
            enfrentamiento = (
                enfrentamientos[i]
                if i < len(enfrentamientos)
                else {
                    "numero": numero_partido,
                    "nombre": f"Partido {numero_partido}",
                    "equipo_a": "",
                    "equipo_b": "",
                }
            )
            equipo_a = enfrentamiento.get("equipo_a", "")
            equipo_b = enfrentamiento.get("equipo_b", "")
            pauta = pauta_etapa[i] if i < len(pauta_etapa) else None
            datos_calendario = calendario_etapa.get(numero_partido, {})
            if not equipo_a:
                equipo_a = datos_calendario.get("equipo_a", "")
            if not equipo_b:
                equipo_b = datos_calendario.get("equipo_b", "")
            nombre_enfrentamiento = enfrentamiento.get("nombre") or f"Partido {numero_partido}"
            if equipo_a or equipo_b:
                nombre_enfrentamiento = formatear_enfrentamiento(equipo_a, equipo_b, numero_partido)

            if cfg["tipo"] == "GRUPOS":
                info_resultado = interpretar_resultado_grupos(pauta, equipo_a, equipo_b)
                pauta_visible = valor_visible(pauta)
            elif cfg["tipo"] == "ELIM":
                info_resultado = interpretar_resultado_eliminatoria(pauta, equipo_a, equipo_b)
                if isinstance(pauta, (list, tuple)):
                    pasa = pauta[0] if len(pauta) > 0 else None
                    modo = pauta[1] if len(pauta) > 1 else None
                    pauta_visible = formatear_prediccion_elim(pasa, modo)
                else:
                    pauta_visible = valor_visible(pauta)
            else:
                info_resultado = {
                    "resultado": "Pendiente",
                    "ganador": "",
                    "winner_side": "",
                    "modo": "",
                    "estado": "pendiente",
                    "outcome": "pending",
                }
                pauta_visible = valor_visible(pauta)

            partidos.append({
                "numero": numero_partido,
                "nombre_enfrentamiento": nombre_enfrentamiento,
                "equipo_a": equipo_a,
                "equipo_b": equipo_b,
                "fecha_chile": datos_calendario.get("fecha_chile", ""),
                "hora_chile": datos_calendario.get("hora_chile", ""),
                "datetime_chile_iso": datos_calendario.get("datetime_chile_iso", ""),
                "fecha_hora_display": datos_calendario.get("fecha_hora_display", "Horario por confirmar"),
                "sort_key": datos_calendario.get("sort_key", ""),
                "sede": datos_calendario.get("sede", ""),
                "notas": datos_calendario.get("notas", ""),
                "pauta": pauta_visible,
                "resultado": info_resultado["resultado"],
                "ganador": info_resultado["ganador"],
                "winner_side": info_resultado["winner_side"],
                "modo": info_resultado["modo"],
                "estado": info_resultado["estado"],
                "outcome": info_resultado["outcome"],
            })

        matches[etapa] = partidos

    return {
        "stages": stages,
        "matches": matches,
    }


def calcular_detalle_etapa(ruta_excel, etapa, pautas_por_etapa):
    cfg = ETAPAS[etapa]
    pauta = pautas_por_etapa.get(etapa)
    if pauta is None:
        raise ValueError(
            f"No existe pauta cargada para la etapa {etapa}. "
            "Revisa la carpeta pauta."
        )

    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    if cfg["tipo"] == "GRUPOS":
        apuestas = leer_celdas_resultado(ws, cfg["n_partidos"])
        if len(pauta) != cfg["n_partidos"]:
            raise ValueError(
                f"La pauta de {etapa} debe tener {cfg['n_partidos']} resultados "
                f"(tiene {len(pauta)})."
            )
        partidos = []
        total_etapa = 0
        for i, (a, r) in enumerate(zip(apuestas, pauta), start=1):
            pauta_llena = normalizar_texto(r) != ""
            puntos_exactitud = 1 if (pauta_llena and normalizar_texto(a) == normalizar_texto(r)) else 0
            puntos_signo = 0
            bonus = 0
            total_partido = puntos_exactitud
            total_etapa += total_partido
            desglose = "Acierta resultado (+1)." if total_partido else "No acierta resultado (0)."
            partidos.append({
                "partido": i,
                "pronostico": valor_visible(a),
                "pauta": valor_visible(r),
                "puntos_exactitud": puntos_exactitud,
                "puntos_signo": puntos_signo,
                "bonus": bonus,
                "total": total_partido,
                "desglose": desglose,
            })
        return {
            "etapa": etapa,
            "tipo": cfg["tipo"],
            "total_etapa": total_etapa,
            "partidos": partidos,
        }

    if cfg["tipo"] == "ELIM":
        apuestas = leer_celdas_eliminatoria(ws, cfg["n_partidos"])
        if len(pauta) != cfg["n_partidos"]:
            raise ValueError(
                f"La pauta de {etapa} debe tener {cfg['n_partidos']} tuplas "
                f"(tiene {len(pauta)})."
            )
        partidos = []
        total_etapa = 0
        for i, ((pasa, modo), (pasa_real, modo_real)) in enumerate(zip(apuestas, pauta), start=1):

            pauta_pasa_llena = normalizar_pasa_eliminatoria(pasa_real) != ""
            pauta_modo_llena = normalizar_texto(modo_real) != ""     
            acierta_pasa = pauta_pasa_llena and (normalizar_pasa_eliminatoria(pasa) == normalizar_pasa_eliminatoria(pasa_real))
            puntos_exactitud = cfg["ppp"] if acierta_pasa else 0
            puntos_signo = 0
            bonus = 1 if (acierta_pasa and pauta_modo_llena and normalizar_texto(modo) == normalizar_texto(modo_real)) else 0
            
            total_partido = puntos_exactitud + bonus
            total_etapa += total_partido
            if not acierta_pasa:
                desglose = "No acierta quién pasa (0)."
            elif bonus:
                desglose = f"Acierta quién pasa (+{cfg['ppp']}) y modo (+1)."
            else:
                desglose = f"Acierta quién pasa (+{cfg['ppp']}), pero falla modo (+0)."

            partidos.append({
                "partido": i,
                "pronostico": formatear_prediccion_elim(pasa, modo),
                "pasa_raw": valor_payload(pasa),
                "modo_raw": valor_payload(modo),
                "pauta": formatear_prediccion_elim(pasa_real, modo_real),
                "puntos_exactitud": puntos_exactitud,
                "puntos_signo": puntos_signo,
                "bonus": bonus,
                "total": total_partido,
                "desglose": desglose,
            })
        return {
            "etapa": etapa,
            "tipo": cfg["tipo"],
            "total_etapa": total_etapa,
            "partidos": partidos,
        }

    raise ValueError(f"Tipo de etapa desconocido: {cfg['tipo']}")


def calcular_puntaje_etapa(ruta_excel, etapa, pautas_por_etapa):
    return calcular_detalle_etapa(ruta_excel, etapa, pautas_por_etapa)["total_etapa"]


def categoria_modo_tendencia(modo):
    norm = normalizar_comparacion(modo)
    if norm in {
        "90", "90 MIN", "90 MINUTOS", "TIEMPO REGULAR",
        "EN LOS 90", "EN 90", "REGULAR",
    }:
        return "90"
    if norm in {
        "120", "120 MIN", "120 MINUTOS", "EN LOS 120", "EN 120",
        "ALARGUE", "TIEMPO EXTRA", "EN TIEMPO EXTRA",
        "SUPLEMENTARIO", "PRORROGA",
    }:
        return "120"
    if norm in {
        "PENALES", "EN PENALES", "PENAL", "PENALTIES",
        "PENALTIS", "PK", "PEN",
    }:
        return "PENALES"
    return None


def normalizar_equipo_simulacion_final(valor):
    """Devuelve el finalista canónico admitido por el simulador."""
    equipo = normalizar_comparacion(normalizar_pasa_eliminatoria(valor))
    return equipo if equipo in {"ESPANA", "ARGENTINA"} else ""


def final_oficial_completa(pautas_por_etapa):
    """Valida exclusivamente si corresponde retirar el simulador de E06."""
    pauta_final = pautas_por_etapa.get("E06") or []
    if not pauta_final:
        return False

    partido_final = pauta_final[0]
    if not pauta_partido_finalizado("E06", partido_final):
        return False

    if isinstance(partido_final, (list, tuple)):
        ganador = partido_final[0] if len(partido_final) > 0 else None
        modo = partido_final[1] if len(partido_final) > 1 else None
    else:
        ganador = partido_final
        modo = None

    return bool(
        normalizar_equipo_simulacion_final(ganador)
        and categoria_modo_tendencia(modo)
    )


def formatear_monto_clp_python(monto):
    """Formatea un monto entero para los mensajes informativos de consola."""
    return "$" + f"{int(monto):,}".replace(",", ".")


def construir_premios_payload(titulo_competencia, pozo_premios,
                               pautas_por_etapa, final_oficial_finalizada=None):
    """Entrega solo configuración; los premios se calculan desde la tabla visible."""
    if final_oficial_finalizada is None:
        final_oficial_finalizada = final_oficial_completa(pautas_por_etapa)

    resultado_oficial = None
    if final_oficial_finalizada:
        partido_final = (pautas_por_etapa.get("E06") or [])[0]
        if isinstance(partido_final, (list, tuple)):
            ganador_raw = partido_final[0] if len(partido_final) > 0 else None
            modo_raw = partido_final[1] if len(partido_final) > 1 else None
        else:
            ganador_raw = partido_final
            modo_raw = None

        ganador = normalizar_equipo_simulacion_final(ganador_raw)
        modo = categoria_modo_tendencia(modo_raw) or ""
        if ganador and modo:
            resultado_oficial = {
                "winner": ganador,
                "mode": modo,
                "winner_label": {
                    "ESPANA": "España",
                    "ARGENTINA": "Argentina",
                }[ganador],
                "mode_label": {
                    "90": "90'",
                    "120": "120'",
                    "PENALES": "penales",
                }[modo],
            }

    return {
        "competition": str(titulo_competencia),
        "prize_pool": max(0, int(pozo_premios or 0)),
        "prize_slots": {
            str(posicion): porcentaje
            for posicion, porcentaje in PREMIOS_POR_POSICION.items()
        },
        "official_final_finished": bool(final_oficial_finalizada),
        "official_final_result": resultado_oficial,
    }


def leer_pronosticos_eliminatoria_crudos(ruta_excel, etapa):
    cfg = ETAPAS[etapa]
    if cfg["tipo"] != "ELIM":
        return []
    wb = load_workbook(ruta_excel, data_only=True)
    try:
        apuestas = leer_celdas_eliminatoria(wb.active, cfg["n_partidos"])
        return [
            {
                "pasa_raw": valor_payload(pasa),
                "modo_raw": valor_payload(modo),
            }
            for pasa, modo in apuestas
        ]
    finally:
        wb.close()


def construir_partidos_clave_eliminatorias(
        enfrentamientos_detalle_por_etapa, pautas_por_etapa,
        calendario_por_etapa=None):
    partidos_config = [
        ("E04", 1, "Cuarto 1"),
        ("E04", 2, "Cuarto 2"),
        ("E04", 3, "Cuarto 3"),
        ("E04", 4, "Cuarto 4"),
        ("E05", 1, "Semi 1"),
        ("E05", 2, "Semi 2"),
        ("E06", 1, "Final"),
    ]
    calendario_por_etapa = calendario_por_etapa or {}
    partidos = []

    for etapa, numero_partido, etiqueta_corta in partidos_config:
        enfrentamientos = enfrentamientos_detalle_por_etapa.get(etapa, []) or []
        enfrentamiento = next(
            (
                item for item in enfrentamientos
                if int(item.get("numero", 0) or 0) == numero_partido
            ),
            {},
        )
        equipo_a = valor_payload(enfrentamiento.get("equipo_a"))
        equipo_b = valor_payload(enfrentamiento.get("equipo_b"))
        enfrentamiento_conocido = bool(equipo_a and equipo_b)
        selector_label = (
            f"{etiqueta_corta} — {equipo_a} vs {equipo_b}"
            if enfrentamiento_conocido
            else f"{etiqueta_corta} — Enfrentamiento por definir"
        )
        calendario = calendario_por_etapa.get(etapa, {}).get(numero_partido, {})
        pauta_etapa = pautas_por_etapa.get(etapa) or []
        pauta_partido = (
            pauta_etapa[numero_partido - 1]
            if numero_partido <= len(pauta_etapa)
            else None
        )
        partidos.append({
            "id": f"{etapa}-{numero_partido}",
            "stage": etapa,
            "match_number": numero_partido,
            "short_label": etiqueta_corta,
            "selector_label": selector_label,
            "equipo_a": equipo_a,
            "equipo_b": equipo_b,
            "datetime_chile_iso": valor_payload(
                calendario.get("datetime_chile_iso")
            ),
            "finished": (
                pauta_partido_finalizado(etapa, pauta_partido)
                if pauta_partido is not None
                else False
            ),
            "matchup_known": enfrentamiento_conocido,
        })

    return partidos


def formatear_pronostico_para_tabla(pasa, modo):
    ganador = texto_pasa_eliminatoria(pasa).strip()
    modo_original = valor_payload(modo)
    ganador_norm = normalizar_comparacion(
        normalizar_pasa_eliminatoria(ganador)
    )
    modo_norm = normalizar_comparacion(modo_original)
    categoria = categoria_modo_tendencia(modo_original)
    etiquetas_modo = {"90": "90'", "120": "120'", "PENALES": "Penales"}

    if not ganador_norm and not modo_norm:
        return {
            "winner": "",
            "mode": "",
            "display": "Sin pronóstico",
            "status": "missing",
        }
    if not ganador_norm:
        return {
            "winner": "",
            "mode": etiquetas_modo.get(categoria, etiqueta_modo_eliminatoria(modo_original)),
            "display": "Pronóstico incompleto",
            "status": "incomplete",
        }
    if not modo_norm:
        return {
            "winner": ganador,
            "mode": "",
            "display": f"{ganador} · Modo no indicado",
            "status": "incomplete",
        }
    if not categoria:
        modo_visible = etiqueta_modo_eliminatoria(modo_original) or modo_original
        return {
            "winner": ganador,
            "mode": modo_visible,
            "display": f"{ganador} · {modo_visible}",
            "status": "unrecognized",
        }

    modo_visible = etiquetas_modo[categoria]
    return {
        "winner": ganador,
        "mode": modo_visible,
        "display": f"{ganador} · {modo_visible}",
        "status": "complete",
    }


def construir_pronosticos_tabla_payload(datos, partidos_clave):
    predictions = {}
    for pid, info in datos.items():
        predictions[pid] = {}
        for partido in partidos_clave:
            etapa = partido["stage"]
            indice = partido["match_number"] - 1
            pronosticos_etapa = info.get("pronosticos_elim", {}).get(etapa, [])
            raw = pronosticos_etapa[indice] if indice < len(pronosticos_etapa) else {}
            pronostico = formatear_pronostico_para_tabla(
                raw.get("pasa_raw", ""), raw.get("modo_raw", "")
            )

            if pronostico["winner"] and partido["matchup_known"]:
                ganador_norm = normalizar_comparacion(
                    normalizar_pasa_eliminatoria(pronostico["winner"])
                )
                equipos_validos = {
                    normalizar_comparacion(partido["equipo_a"]),
                    normalizar_comparacion(partido["equipo_b"]),
                }
                if ganador_norm not in equipos_validos:
                    pronostico["status"] = "unrecognized"

            predictions[pid][partido["id"]] = pronostico

    return {
        "matches": partidos_clave,
        "predictions": predictions,
    }


def construir_simulacion_final_payload(datos, filas_participantes):
    """Construye los datos originales y normalizados para la simulación local."""
    participantes_payload = {}

    for orden_original, fila in enumerate(filas_participantes):
        pid, posicion, nombre, scores, bono, total = fila
        info = datos.get(pid, {})
        pronosticos_final = info.get("pronosticos_elim", {}).get("E06", [])
        pronostico_final = (
            pronosticos_final[0]
            if pronosticos_final and isinstance(pronosticos_final[0], dict)
            else {}
        )
        ganador = normalizar_equipo_simulacion_final(
            pronostico_final.get("pasa_raw", "")
        )
        modo = categoria_modo_tendencia(
            pronostico_final.get("modo_raw", "")
        ) or ""
        campeon_predicho = normalizar_comparacion(info.get("campeon_pred"))
        base_sin_final_ni_bono = sum(
            scores.get(etapa, 0) or 0
            for etapa in ETAPAS
            if etapa != "E06"
        )

        participantes_payload[pid] = {
            "id": pid,
            "name": nombre,
            "sort_name": str(nombre).upper(),
            "base_without_final_and_bonus": base_sin_final_ni_bono,
            "original_final_points": scores.get("E06", 0),
            "original_bonus_points": bono,
            "original_total": total,
            "original_position": posicion,
            "original_order": orden_original,
            "final_prediction": {
                "winner": ganador,
                "mode": modo,
            },
            "champion_prediction": campeon_predicho,
        }

    return {
        "scoring": {
            "winner_points": ETAPAS["E06"]["ppp"],
            "mode_points": 1,
            "champion_bonus_points": BONUS_PTS,
        },
        "participants": participantes_payload,
    }


def construir_tendencias_eliminatorias(datos, partidos_clave):
    total_participantes = len(datos)
    payload_partidos = []

    for partido_base in partidos_clave:
        etapa = partido_base["stage"]
        numero_partido = partido_base["match_number"]
        equipo_a = partido_base["equipo_a"]
        equipo_b = partido_base["equipo_b"]

        conteo_equipos = {"A": 0, "B": 0}
        conteo_matriz = {
            "A": {"90": 0, "120": 0, "PENALES": 0},
            "B": {"90": 0, "120": 0, "PENALES": 0},
        }
        sin_pronostico = 0
        no_reconocidos = 0

        equipo_a_norm = normalizar_comparacion(equipo_a)
        equipo_b_norm = normalizar_comparacion(equipo_b)

        for info in datos.values():
            pronosticos_etapa = info.get("pronosticos_elim", {}).get(etapa, [])
            indice = numero_partido - 1
            if indice >= len(pronosticos_etapa):
                sin_pronostico += 1
                continue

            pronostico = pronosticos_etapa[indice]
            pasa_raw = pronostico.get("pasa_raw", "")
            modo_raw = pronostico.get("modo_raw", "")
            pasa_sin_prefijo = texto_pasa_eliminatoria(pasa_raw)
            pasa_norm = normalizar_comparacion(
                normalizar_pasa_eliminatoria(pasa_sin_prefijo)
            )
            modo_norm = normalizar_comparacion(modo_raw)
            if not pasa_norm and not modo_norm:
                sin_pronostico += 1
                continue

            lado = None
            if equipo_a_norm and pasa_norm == equipo_a_norm:
                lado = "A"
            elif equipo_b_norm and pasa_norm == equipo_b_norm:
                lado = "B"

            categoria_modo = categoria_modo_tendencia(modo_raw)
            if lado:
                conteo_equipos[lado] += 1
            if lado and categoria_modo:
                conteo_matriz[lado][categoria_modo] += 1
            else:
                no_reconocidos += 1

        pronosticos_equipo_validos = sum(conteo_equipos.values())
        pronosticos_combinacion_validos = sum(
            sum(modos.values()) for modos in conteo_matriz.values()
        )

        def porcentaje(conteo, denominador):
            return (conteo / denominador * 100) if denominador else 0

        payload_partido = dict(partido_base)
        payload_partido.update({
            "total_participants": total_participantes,
            "valid_team_predictions": pronosticos_equipo_validos,
            "valid_combination_predictions": pronosticos_combinacion_validos,
            "missing_predictions": sin_pronostico,
            "unrecognized_predictions": no_reconocidos,
            "teams": {
                "A": {
                    "name": equipo_a,
                    "count": conteo_equipos["A"],
                    "percentage": porcentaje(
                        conteo_equipos["A"], pronosticos_equipo_validos
                    ),
                },
                "B": {
                    "name": equipo_b,
                    "count": conteo_equipos["B"],
                    "percentage": porcentaje(
                        conteo_equipos["B"], pronosticos_equipo_validos
                    ),
                },
            },
            "matrix": {
                lado: {
                    modo: {
                        "count": conteo,
                        "percentage": porcentaje(
                            conteo, pronosticos_combinacion_validos
                        ),
                    }
                    for modo, conteo in modos.items()
                }
                for lado, modos in conteo_matriz.items()
            },
        })
        payload_partidos.append(payload_partido)

    return {"matches": payload_partidos}

def html_escape(x):
    if x is None:
        return ""
    return (str(x)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;"))

def calcular_posiciones_con_empate(participantes):
    posiciones = []
    pos_anterior = None
    total_anterior = None

    for idx, participante in enumerate(participantes, start=1):
        total = participante[4]
        if total != total_anterior:
            pos_actual = idx
        else:
            pos_actual = pos_anterior

        posiciones.append(pos_actual)
        total_anterior = total
        pos_anterior = pos_actual

    return posiciones


def formatear_puntaje_familiar(valor):
    numero = float(valor)
    return f"{numero:.1f}"


def construir_ranking_familiar(participantes, datos_familias, etapas_ordenadas):
    mapa_excel = datos_familias["participantes"]
    familias_excel = datos_familias["familias"]
    integrantes_por_familia = {clave: [] for clave in familias_excel}
    nombres_excel_usados = set()
    participantes_sin_familia = []

    for participante in participantes:
        pid, nombre, scores, bono, total, errores = participante
        nombre_norm = normalizar_comparacion(nombre)
        asignacion = mapa_excel.get(nombre_norm)
        if not asignacion:
            participantes_sin_familia.append(nombre)
            continue
        nombres_excel_usados.add(nombre_norm)
        integrantes_por_familia[asignacion["familia_norm"]].append(participante)

    nombres_excel_sin_coincidencia = [
        info["nombre_excel"]
        for nombre_norm, info in mapa_excel.items()
        if nombre_norm not in nombres_excel_usados
    ]
    familias_sin_integrantes = []
    ranking = []
    integrantes_usados = {}

    for familia_norm, familia_info in familias_excel.items():
        integrantes = integrantes_por_familia[familia_norm]
        if not integrantes:
            familias_sin_integrantes.append(familia_info["nombre"])
            continue

        cantidad = len(integrantes)
        scores_promedio = {
            etapa: sum(p[2].get(etapa, 0) for p in integrantes) / cantidad
            for etapa in etapas_ordenadas
        }
        bono_promedio = sum(p[3] for p in integrantes) / cantidad
        total_promedio = sum(p[4] for p in integrantes) / cantidad
        total_por_componentes = sum(scores_promedio.values()) + bono_promedio
        if abs(total_promedio - total_por_componentes) > 1e-9:
            raise ValueError(
                f"el total promedio de la familia '{familia_info['nombre']}' "
                "no coincide con la suma de sus promedios por etapa y bono."
            )

        familia_id = f"familia:{familia_norm}"
        ranking.append((
            familia_id,
            familia_info["nombre"],
            scores_promedio,
            bono_promedio,
            total_promedio,
            [],
        ))
        integrantes_usados[familia_id] = sorted(
            (p[1] for p in integrantes),
            key=normalizar_comparacion,
        )

    ranking.sort(key=lambda x: (-x[4], normalizar_comparacion(x[1])))
    return {
        "ranking": ranking,
        "integrantes": integrantes_usados,
        "participantes_sin_familia": sorted(
            participantes_sin_familia, key=normalizar_comparacion
        ),
        "nombres_excel_sin_coincidencia": sorted(
            nombres_excel_sin_coincidencia, key=normalizar_comparacion
        ),
        "familias_sin_integrantes": sorted(
            familias_sin_integrantes, key=normalizar_comparacion
        ),
        "cantidad_familias_excel": len(familias_excel),
    }


def calcular_puntos_repartidos(pautas_por_etapa, campeon_real_oficial, max_por_etapa, max_bonus):
    puntos_repartidos = 0

    for etapa, cfg in ETAPAS.items():
        pauta = pautas_por_etapa.get(etapa)
        if not pauta:
            continue

        puntos_etapa = 0
        if cfg["tipo"] == "GRUPOS":
            for resultado_real in pauta[:cfg["n_partidos"]]:
                if normalizar_texto(resultado_real) != "":
                    puntos_etapa += cfg["ppp"]
        elif cfg["tipo"] == "ELIM":
            for item in pauta[:cfg["n_partidos"]]:
                if isinstance(item, (list, tuple)):
                    pasa_real = item[0] if len(item) > 0 else None
                    modo_real = item[1] if len(item) > 1 else None
                else:
                    pasa_real = item
                    modo_real = None

                if normalizar_pasa_eliminatoria(pasa_real) == "":
                    continue

                puntos_etapa += cfg["ppp"]
                if normalizar_texto(modo_real) != "":
                    puntos_etapa += 1

        puntos_repartidos += min(puntos_etapa, max_por_etapa.get(etapa, puntos_etapa))

    if normalizar_texto(campeon_real_oficial) != "":
        puntos_repartidos += max_bonus

    return puntos_repartidos


def formatear_porcentaje_avance(porcentaje):
    porcentaje_redondeado = round(float(porcentaje), 1)
    if porcentaje_redondeado.is_integer():
        return f"{int(porcentaje_redondeado)}%"
    return f"{porcentaje_redondeado:.1f}%".replace(".", ",")


def calcular_podios_por_etapa(participantes, etapas_ordenadas, etapas_finalizadas,
                              ranking_familiar=False):
    medallas = {
        1: {"medal": "🥇", "class": "stage-gold"},
        2: {"medal": "🥈", "class": "stage-silver"},
        3: {"medal": "🥉", "class": "stage-bronze"},
    }
    podios_por_etapa = {}

    for etapa in etapas_ordenadas:
        if not etapas_finalizadas.get(etapa):
            continue

        puntajes = []
        for pid, nombre, scores, bono, total, errores in participantes:
            puntaje = scores.get(etapa, 0)
            if puntaje > 0:
                puntajes.append((pid, puntaje))

        if not puntajes:
            continue

        puntajes_distintos = sorted({puntaje for pid, puntaje in puntajes}, reverse=True)
        rank_por_puntaje = {
            puntaje: rank
            for rank, puntaje in enumerate(puntajes_distintos[:3], start=1)
        }

        for pid, puntaje in puntajes:
            rank_actual = rank_por_puntaje.get(puntaje)
            if rank_actual not in medallas:
                continue

            info_medalla = medallas[rank_actual]
            podios_por_etapa.setdefault(pid, {})[etapa] = {
                "rank": rank_actual,
                "medal": info_medalla["medal"],
                "class": info_medalla["class"],
                "title": (
                    f"{rank_actual}° lugar familiar en {ETIQUETAS_ETAPAS.get(etapa, etapa)}"
                    if ranking_familiar
                    else f"{rank_actual}° lugar en {ETIQUETAS_ETAPAS.get(etapa, etapa)}"
                ),
            }

    return podios_por_etapa


def render_tabla_posiciones_html(filas, etapas_ordenadas, max_por_etapa,
                                  max_bonus, max_total, podios_por_etapa=None,
                                  titulo=None, ranking_familiar=False,
                                  mostrar_columna_pronostico=False,
                                  campeones_por_participante=None):
    podios_por_etapa = podios_por_etapa or {}
    campeones_por_participante = campeones_por_participante or {}
    headers = ["Pos", "Nombre", "Total"]
    max_row = ["", "", f"Max={max_total}"]
    if mostrar_columna_pronostico:
        headers.append("Pronóstico")
        headers.append("Bono pronóstico campeón")
        max_row.append("—")
        max_row.append(f"Max={max_bonus}")
    headers += [ETIQUETAS_ETAPAS[e] for e in etapas_ordenadas]
    max_row += [f"Max={max_por_etapa[e]}" for e in etapas_ordenadas]
    if not mostrar_columna_pronostico:
        headers.append(NOMBRE_COLUMNA_BONUS)
        max_row.append(f"Max={max_bonus}")

    colgroup_html = (
        "<colgroup>"
        "<col class='col-pos'>"
        "<col class='col-nombre'>"
        "<col class='col-total'>"
        + ("<col class='col-pronostico'>" if mostrar_columna_pronostico else "")
        + ("<col class='col-campeon'>" if mostrar_columna_pronostico else "")
        + "".join("<col class='col-puntaje'>" for _ in etapas_ordenadas)
        + ("" if mostrar_columna_pronostico else "<col class='col-puntaje'>")
        + "</colgroup>"
    )

    def mostrar(valor):
        return formatear_puntaje_familiar(valor) if ranking_familiar else str(valor)

    def render_header_cell(texto, idx, fila_maximos=False):
        if mostrar_columna_pronostico and idx == 3 and not fila_maximos:
            return (
                "<th class='col-pronostico-header'>"
                "<span>Pronóstico</span>"
                "<small id='pronostico-header-partido'>Partido seleccionado</small>"
                "</th>"
            )
        clase = " class='total'" if idx == 2 else ""
        return f"<th{clase}>{html_escape(texto)}</th>"

    def clase_podio_tabla(pos):
        return {
            1: "podio-oro",
            2: "podio-plata",
            3: "podio-bronce",
        }.get(pos, "")

    def render_stage_score(pid, etapa, puntaje):
        valor_visible = mostrar(puntaje)
        podio = podios_por_etapa.get(pid, {}).get(etapa)
        if not podio:
            return html_escape(valor_visible)
        return (
            f"<span class='stage-score {podio['class']}' "
            f"title='{html_escape(podio['title'])}'>"
            f"<span class='stage-score-value'>{html_escape(valor_visible)}</span>"
            f"<span class='stage-medal' aria-hidden='true'>{podio['medal']}</span>"
            "</span>"
        )

    body_html = []
    for orden_original, (pid, pos, nombre, scores, bono, total) in enumerate(filas):
        clase = clase_podio_tabla(pos)
        atributos = f" data-participant-id='{html_escape(pid)}'"
        if mostrar_columna_pronostico:
            atributos += (
                f" data-participant-name='{html_escape(nombre)}'"
                f" data-original-order='{orden_original}'"
            )
        row_open = (
            f"<tr class='{clase}'{atributos}>" if clase else f"<tr{atributos}>"
        )
        if mostrar_columna_pronostico:
            cells = [
                "<td class='simulacion-pos-cell' "
                f"data-original-value='{html_escape(pos)}'>{html_escape(pos)}</td>",
                f"<td class='nombre'>{html_escape(nombre)}</td>",
                "<td class='total simulacion-total-cell' "
                f"data-original-value='{html_escape(mostrar(total))}'>"
                f"{html_escape(mostrar(total))}</td>",
            ]
        else:
            cells = [
                f"<td>{html_escape(pos)}</td>",
                f"<td class='nombre'>{html_escape(nombre)}</td>",
                f"<td class='total'>{html_escape(mostrar(total))}</td>",
            ]
        if mostrar_columna_pronostico:
            cells.append(
                "<td class='pronostico-partido-cell' "
                f"data-participant-id='{html_escape(pid)}'>Sin pronóstico</td>"
            )
            campeon = valor_payload(campeones_por_participante.get(pid))
            bono_visible = html_escape(mostrar(bono))
            if campeon:
                chip_campeon = (
                    "<span class='pronostico-chip pronostico-chip-complete "
                    f"pronostico-campeon-chip' data-country='{html_escape(campeon)}'>"
                    "<span class='pronostico-campeon-nombre'>"
                    f"{html_escape(campeon)}</span>"
                    "<span class='pronostico-campeon-separador' aria-hidden='true'>·</span>"
                    "<span class='pronostico-campeon-puntos' "
                    f"data-original-value='{bono_visible}'>"
                    f"{bono_visible}</span></span>"
                )
            else:
                chip_campeon = (
                    "<span class='pronostico-chip pronostico-chip-missing'>"
                    "<span class='pronostico-campeon-nombre'>Sin pronóstico</span>"
                    "<span class='pronostico-campeon-separador' aria-hidden='true'>·</span>"
                    "<span class='pronostico-campeon-puntos' "
                    f"data-original-value='{bono_visible}'>"
                    f"{bono_visible}</span></span>"
                )
            cells.append(
                f"<td class='pronostico-campeon-cell'>{chip_campeon}</td>"
            )
        for etapa in etapas_ordenadas:
            atributos_etapa = (
                " class='simulacion-final-cell' "
                f"data-original-value='{html_escape(mostrar(scores.get(etapa, 0)))}'"
                if mostrar_columna_pronostico and etapa == "E06"
                else ""
            )
            cells.append(
                f"<td{atributos_etapa}>"
                f"{render_stage_score(pid, etapa, scores.get(etapa, 0))}</td>"
            )
        if not mostrar_columna_pronostico:
            cells.append(f"<td>{html_escape(mostrar(bono))}</td>")
        body_html.append(row_open + "".join(cells) + "</tr>")

    titulo_html = f"<h2 class='ranking-title'>{html_escape(titulo)}</h2>\n" if titulo else ""
    tipo_tabla = "tabla-individual" if mostrar_columna_pronostico else "tabla-familiar"
    return f"""{titulo_html}<div class="tabla-posiciones-scroll {tipo_tabla}-scroll">
<table class="tabla-posiciones {tipo_tabla}">
{colgroup_html}
<thead>
<tr>
{''.join(render_header_cell(h, i) for i, h in enumerate(headers))}
</tr>
<tr>
{''.join(render_header_cell(v, i, True) for i, v in enumerate(max_row))}
</tr>
</thead>
<tbody>
{''.join(body_html)}
</tbody>
</table>
</div>"""


def render_tabla_html(nombre_competencia, participantes, etapas_ordenadas,
                      max_por_etapa, max_bonus, max_total, out_path,
                      detalle_payload, resultados_payload=None,
                      tendencias_payload=None,
                      pronosticos_tabla_payload=None,
                      puntos_repartidos=0, porcentaje_avance=0,
                       podios_por_etapa=None, participantes_familiares=None,
                       podios_familiares=None,
                       campeones_por_participante=None,
                       mostrar_simulador_final=False,
                       simulacion_final_payload=None,
                       pozo_premios=0,
                       premios_payload=None):

    now = datetime.now(ZoneInfo("America/Santiago")).strftime("%Y-%m-%d %H:%M:%S")
    titulo_competencia = html_escape(nombre_competencia)
    detalle_json = json.dumps(detalle_payload, ensure_ascii=False).replace("</", "<\\/")
    resultados_payload = resultados_payload or {"stages": [], "matches": {}}
    resultados_json = json.dumps(resultados_payload, ensure_ascii=False).replace("</", "<\\/")
    tendencias_payload = tendencias_payload or {"matches": []}
    tendencias_json = json.dumps(tendencias_payload, ensure_ascii=False).replace("</", "<\\/")
    pronosticos_tabla_payload = pronosticos_tabla_payload or {
        "matches": [], "predictions": {}
    }
    pronosticos_tabla_json = json.dumps(
        pronosticos_tabla_payload, ensure_ascii=False
    ).replace("</", "<\\/")
    simulacion_final_json = ""
    if mostrar_simulador_final:
        simulacion_final_json = json.dumps(
            simulacion_final_payload or {"scoring": {}, "participants": {}},
            ensure_ascii=False,
        ).replace("</", "<\\/")
    premios_config = {
        "competition": str(nombre_competencia),
        "prize_pool": max(0, int(pozo_premios or 0)),
        "prize_slots": {
            str(posicion): porcentaje
            for posicion, porcentaje in PREMIOS_POR_POSICION.items()
        },
        "official_final_finished": not mostrar_simulador_final,
        "official_final_result": None,
    }
    if premios_payload:
        premios_config.update(premios_payload)
    premios_config["prize_pool"] = max(0, int(pozo_premios or 0))
    premios_json = json.dumps(
        premios_config, ensure_ascii=False
    ).replace("</", "<\\/")
    porcentaje_display = formatear_porcentaje_avance(porcentaje_avance)
    progreso_width = max(0, min(100, float(porcentaje_avance or 0)))
    podios_por_etapa = podios_por_etapa or {}

    mostrar_ranking_familiar = participantes_familiares is not None
    tabla_individual_html = render_tabla_posiciones_html(
        filas=participantes,
        etapas_ordenadas=etapas_ordenadas,
        max_por_etapa=max_por_etapa,
        max_bonus=max_bonus,
        max_total=max_total,
        podios_por_etapa=podios_por_etapa,
        titulo="Tabla individual" if mostrar_ranking_familiar else None,
        mostrar_columna_pronostico=True,
        campeones_por_participante=campeones_por_participante,
    )
    ranking_toggle_html = ""
    ranking_familiar_css = ""
    ranking_toggle_script = ""
    rankings_html = tabla_individual_html
    pronosticos_control_html = """
<div id="pronosticos-tabla-control" class="pronosticos-tabla-control">
    <label for="pronosticos-tabla-selector">Pronósticos mostrados</label>
    <select id="pronosticos-tabla-selector" class="pronosticos-tabla-selector"></select>
</div>
"""
    simulacion_final_control_html = ""
    simulacion_final_data_html = ""
    if mostrar_simulador_final:
        simulacion_final_control_html = """
<section id="simulacion-final-control" class="simulacion-final-control"
         aria-labelledby="simulacion-final-titulo">
    <div class="simulacion-final-copy">
        <h2 id="simulacion-final-titulo">Simular resultado de la final</h2>
        <p id="simulacion-final-subtitulo">
            Selecciona un resultado para ver cómo terminaría la tabla
        </p>
    </div>
    <div class="simulacion-final-field">
        <label for="simulacion-final-selector">Resultado de España vs Argentina</label>
        <div class="simulacion-final-selector-row">
            <select id="simulacion-final-selector" class="simulacion-final-selector"
                    aria-describedby="simulacion-final-subtitulo">
                <option value="" data-scenario-label="tabla actual, sin simulación">
                    Tabla actual — sin simulación
                </option>
                <option value="ESPANA|90" data-winner="ESPANA" data-mode="90"
                        data-winner-label="España" data-mode-label="90 minutos"
                        data-scenario-label="España campeón en 90'">España · 90'</option>
                <option value="ESPANA|120" data-winner="ESPANA" data-mode="120"
                        data-winner-label="España" data-mode-label="120 minutos"
                        data-scenario-label="España campeón en 120'">España · 120'</option>
                <option value="ESPANA|PENALES" data-winner="ESPANA" data-mode="PENALES"
                        data-winner-label="España" data-mode-label="penales"
                        data-scenario-label="España campeón en penales">España · Penales</option>
                <option value="ARGENTINA|90" data-winner="ARGENTINA" data-mode="90"
                        data-winner-label="Argentina" data-mode-label="90 minutos"
                        data-scenario-label="Argentina campeón en 90'">Argentina · 90'</option>
                <option value="ARGENTINA|120" data-winner="ARGENTINA" data-mode="120"
                        data-winner-label="Argentina" data-mode-label="120 minutos"
                        data-scenario-label="Argentina campeón en 120'">Argentina · 120'</option>
                <option value="ARGENTINA|PENALES" data-winner="ARGENTINA" data-mode="PENALES"
                        data-winner-label="Argentina" data-mode-label="penales"
                        data-scenario-label="Argentina campeón en penales">Argentina · Penales</option>
            </select>
            <span id="simulacion-final-badge" class="simulacion-final-badge" hidden>
                Simulación local
            </span>
        </div>
        <p id="simulacion-final-live" class="simulacion-final-live" aria-live="polite"></p>
    </div>
</section>
"""
        simulacion_final_data_html = (
            '<script id="simulacion-final-data" type="application/json">'
            + simulacion_final_json
            + "</script>"
        )

    clase_acciones = " tabla-acciones-solo-premios" if not mostrar_simulador_final else ""
    tabla_acciones_individuales_html = f"""
<div id="tabla-acciones-individuales"
     class="tabla-acciones-individuales{clase_acciones}">
    {simulacion_final_control_html}
    <button type="button" id="abrir-premios-button" class="premios-button"
            aria-haspopup="dialog" aria-controls="premios-dialog">
        <span aria-hidden="true">🏆</span>
        <span>Ver podio y premios</span>
    </button>
</div>
"""
    premios_data_html = (
        '<script id="premios-data" type="application/json">'
        + premios_json
        + "</script>"
    )
    premios_dialog_html = """
<dialog id="premios-dialog" class="premios-dialog"
        aria-labelledby="premios-dialog-title">
    <div class="premios-dialog-content" role="document">
        <button type="button" class="premios-dialog-close premios-dialog-close-x"
                aria-label="Cerrar ventana de premios">×</button>
        <header class="premios-dialog-header">
            <p class="premios-dialog-kicker">Premios de la competencia</p>
            <h2 id="premios-dialog-title">Podio y distribución de premios</h2>
        </header>
        <div class="premios-resumen">
            <div>
                <span class="premios-resumen-label">Pozo total:</span>
                <strong id="premios-pozo"></strong>
            </div>
        </div>
        <div id="premios-podio" class="premios-podio"
             aria-live="polite" aria-atomic="true"></div>
        <p id="premios-redondeo-nota" class="premios-redondeo-nota" hidden></p>
        <div class="premios-dialog-footer">
            <button type="button"
                    class="premios-dialog-close premios-dialog-close-button">Cerrar</button>
        </div>
    </div>
</dialog>
"""

    if mostrar_ranking_familiar:
        tabla_familiar_html = render_tabla_posiciones_html(
            filas=participantes_familiares,
            etapas_ordenadas=etapas_ordenadas,
            max_por_etapa=max_por_etapa,
            max_bonus=max_bonus,
            max_total=max_total,
            podios_por_etapa=podios_familiares,
            titulo="Tabla familiar",
            ranking_familiar=True,
            mostrar_columna_pronostico=False,
            campeones_por_participante=None,
        )
        ranking_toggle_html = """
<div class="ranking-toggle" role="group" aria-label="Modalidad de la tabla de posiciones">
    <button type="button" class="ranking-toggle-button is-active"
            data-ranking-target="individual" aria-pressed="true"
            aria-controls="ranking-individual">Individual</button>
    <button type="button" class="ranking-toggle-button"
            data-ranking-target="familiar" aria-pressed="false"
            aria-controls="ranking-familiar">Familiar</button>
</div>
"""
        rankings_html = f"""
<div id="ranking-individual" class="ranking-panel">
{tabla_individual_html}
</div>
<div id="ranking-familiar" class="ranking-panel" hidden>
{tabla_familiar_html}
</div>
"""
        ranking_familiar_css = """
.ranking-toggle {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    width: min(100%, 360px);
    margin: 0 auto 22px;
    padding: 5px;
    border: 1px solid rgba(255, 255, 255, 0.14);
    border-radius: 14px;
    background: rgba(7, 11, 23, 0.72);
    box-shadow: 0 10px 24px rgba(0, 0, 0, 0.22);
}
.ranking-toggle-button {
    min-height: 42px;
    padding: 9px 18px;
    border: 0;
    border-radius: 10px;
    background: transparent;
    color: var(--muted);
    font: inherit;
    font-weight: 800;
    cursor: pointer;
    transition: background-color 160ms ease, color 160ms ease, box-shadow 160ms ease;
}
.ranking-toggle-button:hover {
    color: #ffffff;
    background: rgba(255, 255, 255, 0.07);
}
.ranking-toggle-button.is-active {
    color: #07101f;
    background: linear-gradient(135deg, #f3b000, #ffdf57);
    box-shadow: 0 6px 16px rgba(243, 176, 0, 0.25);
}
.ranking-toggle-button:focus-visible {
    outline: 3px solid #8cc8ff;
    outline-offset: 2px;
}
.ranking-title {
    margin: 0 0 14px;
    text-align: center;
    font-size: clamp(20px, 3vw, 26px);
}
.ranking-panel[hidden] {
    display: none;
}
@media (max-width: 600px) {
    .ranking-toggle {
        width: 100%;
        box-sizing: border-box;
    }
    .ranking-toggle-button {
        min-height: 44px;
        padding-left: 10px;
        padding-right: 10px;
    }
}
"""
        ranking_toggle_script = """
<script>
(function () {
    var buttons = document.querySelectorAll(".ranking-toggle-button");
    var individual = document.getElementById("ranking-individual");
    var familiar = document.getElementById("ranking-familiar");
    var pronosticosControl = document.getElementById("pronosticos-tabla-control");
    var accionesIndividuales = document.getElementById("tabla-acciones-individuales");
    if (!buttons.length || !individual || !familiar) return;

    buttons.forEach(function (button) {
        button.addEventListener("click", function () {
            var mostrarFamiliar = button.getAttribute("data-ranking-target") === "familiar";
            individual.hidden = mostrarFamiliar;
            familiar.hidden = !mostrarFamiliar;
            if (pronosticosControl) pronosticosControl.hidden = mostrarFamiliar;
            if (accionesIndividuales) accionesIndividuales.hidden = mostrarFamiliar;
            buttons.forEach(function (item) {
                var activo = item === button;
                item.classList.toggle("is-active", activo);
                item.setAttribute("aria-pressed", activo ? "true" : "false");
            });
        });
    });
})();
</script>
"""

    detalle_script = """
<script id="resultados-data" type="application/json">__RESULTADOS_JSON__</script>
<script id="detalle-data" type="application/json">__DETALLE_JSON__</script>
<script>
(function () {
    var dataNode = document.getElementById("resultados-data");
    var stageSelect = document.getElementById("resultados-etapa");
    var stageLabel = document.getElementById("resultados-stage-label");
    var stageCount = document.getElementById("resultados-stage-count");
    var tableWrap = document.getElementById("resultados-table-wrap");
    var tableBody = document.getElementById("resultados-body");
    var emptyBox = document.getElementById("resultados-empty");
    var noteBox = document.getElementById("resultados-note");

    if (!dataNode || !stageSelect || !tableWrap || !tableBody || !emptyBox) return;

    var payload = {};
    try {
        payload = JSON.parse(dataNode.textContent || "{}");
    } catch (e) {
        console.error("No pude parsear resultados-data", e);
        return;
    }

    var stages = payload.stages || [];
    var matchesByStage = payload.matches || {};
    var FLAG_BY_NAME = {
        "ALEMANIA": "🇩🇪",
        "ARGENTINA": "🇦🇷",
        "ARABIA SAUDITA": "🇸🇦",
        "ARGELIA": "🇩🇿",
        "AUSTRIA": "🇦🇹",
        "AUSTRALIA": "🇦🇺",
        "BELGICA": "🇧🇪",
        "BOLIVIA": "🇧🇴",
        "BOSNIA": "🇧🇦",
        "BOSNIA AND HERZEGOVINA": "🇧🇦",
        "BOSNIA HERZAGOBINA": "🇧🇦",
        "BOSNIA HERZEGOVINA": "🇧🇦",
        "BOSNIA Y HERZAGOBINA": "🇧🇦",
        "BOSNIA Y HERZEGOVINA": "🇧🇦",
        "BRASIL": "🇧🇷",
        "CABO VERDE": "🇨🇻",
        "CAMERUN": "🇨🇲",
        "CANADA": "🇨🇦",
        "CHILE": "🇨🇱",
        "CHINA": "🇨🇳",
        "COLOMBIA": "🇨🇴",
        "CONGO": "🇨🇬",
        "COREA": "🇰🇷",
        "COREA DEL SUR": "🇰🇷",
        "COSTA DE MARFIL": "🇨🇮",
        "COSTA RICA": "🇨🇷",
        "CROACIA": "🇭🇷",
        "DINAMARCA": "🇩🇰",
        "ECUADOR": "🇪🇨",
        "EGIPTO": "🇪🇬",
        "EE UU": "🇺🇸",
        "EEUU": "🇺🇸",
        "ESCOCIA": "🏴",
        "ESPANA": "🇪🇸",
        "ESTADOS UNIDOS": "🇺🇸",
        "ESTADOS UNIDOS DE AMERICA": "🇺🇸",
        "FRANCIA": "🇫🇷",
        "GALES": "🏴",
        "GHANA": "🇬🇭",
        "HOLANDA": "🇳🇱",
        "INGLATERRA": "🏴",
        "IRAN": "🇮🇷",
        "ITALIA": "🇮🇹",
        "JAPON": "🇯🇵",
        "MARRUECOS": "🇲🇦",
        "MEXICO": "🇲🇽",
        "NIGERIA": "🇳🇬",
        "NORUEGA": "🇳🇴",
        "PAISES BAJOS": "🇳🇱",
        "PANAMA": "🇵🇦",
        "PARAGUAY": "🇵🇾",
        "PERU": "🇵🇪",
        "POLONIA": "🇵🇱",
        "PORTUGAL": "🇵🇹",
        "QATAR": "🇶🇦",
        "CONGO DEMOCRATICO": "🇨🇩",
        "CONGO KINSHASA": "🇨🇩",
        "DR CONGO": "🇨🇩",
        "R D CONGO": "🇨🇩",
        "RD CONGO": "🇨🇩",
        "REP DEMOCRATICA DEL CONGO": "🇨🇩",
        "REPUBLICA DEMOCRATICA DEL CONGO": "🇨🇩",
        "SENEGAL": "🇸🇳",
        "SERBIA": "🇷🇸",
        "SUDAFRICA": "🇿🇦",
        "SUECIA": "🇸🇪",
        "SUIZA": "🇨🇭",
        "TUNEZ": "🇹🇳",
        "TURQUIA": "🇹🇷",
        "URUGUAY": "🇺🇾",
        "USA": "🇺🇸",
        "VENEZUELA": "🇻🇪"
    };
    var COUNTRY_CODE_BY_NAME = {
        "ALEMANIA": "de",
        "ARGENTINA": "ar",
        "ARABIA SAUDITA": "sa",
        "ARGELIA": "dz",
        "AUSTRIA": "at",
        "AUSTRALIA": "au",
        "BELGICA": "be",
        "BOLIVIA": "bo",
        "BOSNIA": "ba",
        "BOSNIA AND HERZEGOVINA": "ba",
        "BOSNIA HERZAGOBINA": "ba",
        "BOSNIA HERZEGOVINA": "ba",
        "BOSNIA Y HERZAGOBINA": "ba",
        "BOSNIA Y HERZEGOVINA": "ba",
        "BRASIL": "br",
        "CABO VERDE": "cv",
        "CAMERUN": "cm",
        "CANADA": "ca",
        "CHILE": "cl",
        "CHINA": "cn",
        "COLOMBIA": "co",
        "CONGO": "cg",
        "COREA": "kr",
        "COREA DEL SUR": "kr",
        "COSTA DE MARFIL": "ci",
        "COSTA RICA": "cr",
        "CROACIA": "hr",
        "DINAMARCA": "dk",
        "ECUADOR": "ec",
        "EGIPTO": "eg",
        "EE UU": "us",
        "EEUU": "us",
        "ESCOCIA": "gb-sct",
        "ESPANA": "es",
        "ESTADOS UNIDOS": "us",
        "ESTADOS UNIDOS DE AMERICA": "us",
        "FRANCIA": "fr",
        "GALES": "gb-wls",
        "GHANA": "gh",
        "HOLANDA": "nl",
        "INGLATERRA": "gb-eng",
        "IRAN": "ir",
        "ITALIA": "it",
        "JAPON": "jp",
        "MARRUECOS": "ma",
        "MEXICO": "mx",
        "NIGERIA": "ng",
        "NORUEGA": "no",
        "PAISES BAJOS": "nl",
        "PANAMA": "pa",
        "PARAGUAY": "py",
        "PERU": "pe",
        "POLONIA": "pl",
        "PORTUGAL": "pt",
        "QATAR": "qa",
        "CONGO DEMOCRATICO": "cd",
        "CONGO KINSHASA": "cd",
        "DR CONGO": "cd",
        "R D CONGO": "cd",
        "RD CONGO": "cd",
        "REP DEMOCRATICA DEL CONGO": "cd",
        "REPUBLICA DEMOCRATICA DEL CONGO": "cd",
        "SENEGAL": "sn",
        "SERBIA": "rs",
        "SUDAFRICA": "za",
        "SUECIA": "se",
        "SUIZA": "ch",
        "TUNEZ": "tn",
        "TURQUIA": "tr",
        "URUGUAY": "uy",
        "USA": "us",
        "VENEZUELA": "ve"
    };

    function normalizarPais(valor) {
        var texto = String(valor || "").trim().toUpperCase();
        if (texto.normalize) {
            texto = texto.normalize("NFD").replace(/[\\u0300-\\u036f]/g, "");
        }
        return texto.replace(/[^A-Z0-9]+/g, " ").replace(/\\s+/g, " ").trim();
    }

    function flagPais(nombre) {
        return FLAG_BY_NAME[normalizarPais(nombre)] || "";
    }

    function codigoPais(nombre) {
        return COUNTRY_CODE_BY_NAME[normalizarPais(nombre)] || "";
    }

    function crearBandera(nombre) {
        var code = codigoPais(nombre);
        var emoji = flagPais(nombre);
        if (code) {
            var img = document.createElement("img");
            img.className = "resultados-flag-img";
            img.src = "https://flagcdn.com/24x18/" + code + ".png";
            img.srcset = "https://flagcdn.com/48x36/" + code + ".png 2x";
            img.alt = emoji || code.toUpperCase();
            img.loading = "lazy";
            img.decoding = "async";
            img.onerror = function () {
                var fallback = document.createElement("span");
                fallback.className = "resultados-flag";
                fallback.textContent = emoji || code.toUpperCase();
                img.replaceWith(fallback);
            };
            return img;
        }
        if (emoji) {
            var flagNode = document.createElement("span");
            flagNode.className = "resultados-flag";
            flagNode.textContent = emoji;
            return flagNode;
        }
        return null;
    }

    window.MundialUI = {
        normalizarPais: normalizarPais,
        crearBandera: crearBandera,
        codigoPais: codigoPais,
        flagPais: flagPais
    };

    function crearPais(nombre, ganador, fallback) {
        var limpio = String(nombre || "").trim();
        var texto = limpio || fallback || "Por definir";
        var span = document.createElement("span");
        span.className = "resultados-team" + (ganador ? " resultados-team-win" : "") + (!limpio ? " resultados-team-pending" : "");

        var flagNode = crearBandera(limpio);
        if (flagNode) {
            span.appendChild(flagNode);
        }

        var nameNode = document.createElement("span");
        nameNode.textContent = texto;
        span.appendChild(nameNode);
        return span;
    }

    function appendCell(tr, className) {
        var td = document.createElement("td");
        if (className) td.className = className;
        tr.appendChild(td);
        return td;
    }

    function buscarEtapaResultado(etapaId) {
        return stages.find(function (stage) { return stage.id === etapaId; }) || null;
    }

    function renderResultadoCell(td, match, stage) {
        var estado = match.estado || "pendiente";
        var outcome = match.outcome || "pending";
        var badge = document.createElement("span");
        badge.className = "resultados-badge resultados-badge-" + (estado === "pendiente" ? "pending" : outcome);
        badge.textContent = estado === "pendiente" ? "Pendiente" : (outcome === "draw" ? "Empate" : "Jugado");
        td.appendChild(badge);

        var line = document.createElement("div");
        line.className = "resultados-result-line";
        if (estado === "pendiente") {
            line.textContent = "Pendiente";
        } else if (outcome === "draw") {
            line.textContent = "Empate";
        } else if (outcome === "winner" && match.ganador) {
            var prefix = document.createElement("span");
            prefix.textContent = stage && stage.type === "ELIM" ? "Pasa " : "Gana ";
            line.appendChild(prefix);
            line.appendChild(crearPais(match.ganador, true, match.ganador));
            if (match.modo) {
                var modo = document.createElement("span");
                modo.className = "resultados-mode";
                modo.textContent = " " + match.modo;
                line.appendChild(modo);
            }
        } else {
            line.textContent = match.resultado || "Resultado";
        }
        td.appendChild(line);
    }

    function renderResultados() {
        var etapaId = stageSelect.value;
        var stage = buscarEtapaResultado(etapaId);
        var partidos = matchesByStage[etapaId] || [];
        tableBody.innerHTML = "";

        if (stageLabel) stageLabel.textContent = stage ? stage.label : "Etapa";
        if (stageCount) {
            stageCount.textContent = partidos.length === 1 ? "1 partido" : String(partidos.length) + " partidos";
        }

        if (!partidos.length) {
            tableWrap.hidden = true;
            if (noteBox) noteBox.hidden = true;
            emptyBox.hidden = false;
            emptyBox.textContent = "Todavía no hay resultados cargados para esta etapa.";
            return;
        }

        var hayJugados = partidos.some(function (match) { return match.estado === "jugado"; });
        emptyBox.hidden = true;
        if (noteBox) {
            noteBox.hidden = hayJugados;
            noteBox.textContent = "Todavía no hay resultados cargados para esta etapa.";
        }

        var partidosOrdenados = partidos.slice().sort(function (a, b) {
            var aTime = Date.parse(String(a.sort_key || "").trim());
            var bTime = Date.parse(String(b.sort_key || "").trim());
            var aTieneHorario = !isNaN(aTime);
            var bTieneHorario = !isNaN(bTime);
            var aNumero = Number(a.numero || 0);
            var bNumero = Number(b.numero || 0);

            if (aTieneHorario && bTieneHorario) {
                if (aTime !== bTime) return aTime - bTime;
                return aNumero - bNumero;
            }
            if (aTieneHorario) return -1;
            if (bTieneHorario) return 1;
            return aNumero - bNumero;
        });

        partidosOrdenados.forEach(function (match) {
            var tr = document.createElement("tr");
            tr.className = "resultados-row resultados-" + (match.estado || "pendiente");

            var partidoCell = appendCell(tr, "resultados-match-cell");
            var numero = document.createElement("strong");
            numero.textContent = "Partido " + String(match.numero || "");
            partidoCell.appendChild(numero);
            var horario = document.createElement("span");
            horario.className = "resultados-schedule-line";
            horario.textContent = match.fecha_hora_display || "Horario por confirmar";
            partidoCell.appendChild(horario);
            var sede = String(match.sede || "").trim();
            if (sede) {
                var sedeNode = document.createElement("span");
                sedeNode.className = "resultados-venue-line";
                sedeNode.textContent = sede;
                partidoCell.appendChild(sedeNode);
            }

            appendCell(tr, "").appendChild(crearPais(match.equipo_a, match.winner_side === "A", "Por definir"));
            appendCell(tr, "").appendChild(crearPais(match.equipo_b, match.winner_side === "B", "Por definir"));
            renderResultadoCell(appendCell(tr, "resultados-result-cell"), match, stage);

            tableBody.appendChild(tr);
        });

        tableWrap.hidden = false;
    }

    function llenarSelectorResultados() {
        stageSelect.innerHTML = "";
        if (!stages.length) {
            var emptyOption = document.createElement("option");
            emptyOption.value = "";
            emptyOption.textContent = "Sin etapas disponibles";
            stageSelect.appendChild(emptyOption);
            stageSelect.disabled = true;
            renderResultados();
            return;
        }

        stages.forEach(function (stage) {
            var option = document.createElement("option");
            option.value = stage.id;
            option.textContent = stage.label;
            stageSelect.appendChild(option);
        });
        stageSelect.disabled = false;
        var etapaConHorario = stages.slice().reverse().find(function (stage) {
            return (matchesByStage[stage.id] || []).some(function (match) {
                var sortKey = String(match.sort_key || "").trim();
                var fecha = String(match.fecha_chile || "").trim();
                var hora = String(match.hora_chile || "").trim();
                return !!sortKey || (!!fecha && !!hora);
            });
        });
        stageSelect.value = (etapaConHorario || stages[0]).id;
        renderResultados();
    }

    stageSelect.addEventListener("change", renderResultados);
    llenarSelectorResultados();
})();

(function () {
    var dataNode = document.getElementById("detalle-data");
    if (!dataNode) return;

    var payload = {};
    try {
        payload = JSON.parse(dataNode.textContent || "{}");
    } catch (e) {
        console.error("No pude parsear detalle-data", e);
        return;
    }

    var participantes = payload.participants || [];
    var etapas = payload.stages || [];
    var detalles = payload.details || {};
    var etiquetasPartidos = payload.match_labels || {};
    var MENSAJE_RONDA_NO_COMIENZA = "Ronda aún no comienza.";

    var participanteSel = document.getElementById("detalle-participante");
    var etapaSel = document.getElementById("detalle-etapa");
    var content = document.getElementById("detalle-content");
    var resumenParticipante = document.getElementById("detalle-resumen-participante");
    var resumenEtapa = document.getElementById("detalle-resumen-etapa");
    var resumenTotal = document.getElementById("detalle-resumen-total");
    var detalleHead = document.getElementById("detalle-head-row");
    var detalleBody = document.getElementById("detalle-body");

    var detalle2EtapaSel = document.getElementById("detalle2-etapa");
    var detalle2PartidoSel = document.getElementById("detalle2-partido");
    var detalle2Content = document.getElementById("detalle2-content");
    var detalle2ResumenEtapa = document.getElementById("detalle2-resumen-etapa");
    var detalle2ResumenPartido = document.getElementById("detalle2-resumen-partido");
    var detalle2ResumenResultado = document.getElementById("detalle2-resumen-resultado");
    var detalle2Head = document.getElementById("detalle2-head-row");
    var detalle2Body = document.getElementById("detalle2-body");

    function appendCell(tr, value, className) {
        var td = document.createElement("td");
        if (className) td.className = className;
        td.textContent = value;
        tr.appendChild(td);
    }

    function clasePuntos(value) {
        return "num " + (Number(value || 0) > 0 ? "pts-pos" : "pts-zero");
    }

    function resetSelect(selectNode, placeholder) {
        if (!selectNode) return;
        selectNode.innerHTML = "";
        var option = document.createElement("option");
        option.value = "";
        option.textContent = placeholder;
        selectNode.appendChild(option);
    }

    function renderHeaderDetalle(mostrarBonus) {
        detalleHead.innerHTML = "";
        var headers = [
            { text: "Partido", cls: "" },
            { text: "Pronóstico participante", cls: "" },
            { text: "Resultado Real", cls: "" },
            { text: "Puntos por acertar resultado", cls: "num" }
        ];

        if (mostrarBonus) {
            headers.push({ text: "Bonus por acertar modo", cls: "num" });
        }
        headers.push({ text: "Total", cls: "num" });

        headers.forEach(function (h) {
            var th = document.createElement("th");
            if (h.cls) th.className = h.cls;
            th.textContent = h.text;
            detalleHead.appendChild(th);
        });
    }

    function renderHeaderPartido(mostrarBonus) {
        detalle2Head.innerHTML = "";
        var headers = [
            { text: "Participante", cls: "" },
            { text: "Pronóstico participante", cls: "" },
            { text: "Resultado Real", cls: "" },
            { text: "Puntos por acertar resultado", cls: "num" }
        ];

        if (mostrarBonus) {
            headers.push({ text: "Bonus por acertar modo", cls: "num" });
        }
        headers.push({ text: "Total partido", cls: "num" });

        headers.forEach(function (h) {
            var th = document.createElement("th");
            if (h.cls) th.className = h.cls;
            th.textContent = h.text;
            detalle2Head.appendChild(th);
        });
    }

    function renderSinDataDetalle(mostrarBonus, mensaje) {
        detalleBody.innerHTML = "";
        var tr = document.createElement("tr");
        var td = document.createElement("td");
        td.colSpan = mostrarBonus ? 6 : 5;
        td.textContent = mensaje;
        tr.appendChild(td);
        detalleBody.appendChild(tr);
    }

    function renderSinDataPartido(mostrarBonus, mensaje) {
        detalle2Body.innerHTML = "";
        var tr = document.createElement("tr");
        var td = document.createElement("td");
        td.colSpan = mostrarBonus ? 6 : 5;
        td.textContent = mensaje;
        tr.appendChild(td);
        detalle2Body.appendChild(tr);
    }

    function buscarEtapa(etapaId) {
        return etapas.find(function (e) { return e.id === etapaId; }) || null;
    }

    function etapaEstaComenzada(etapa) {
        return !!(etapa && etapa.started);
    }

    function etiquetaPartido(etapaId, numeroPartido) {
        var n = Number(numeroPartido);
        if (isNaN(n) || n <= 0) {
            return "";
        }
        var lista = etiquetasPartidos[etapaId];
        if (Array.isArray(lista) && (n - 1) < lista.length) {
            var etiqueta = String(lista[n - 1] || "").trim();
            if (etiqueta) {
                return etiqueta;
            }
        }
        return "Partido " + String(n);
    }

    function partidosDisponiblesDeEtapa(etapaId) {
        var partidos = {};
        participantes.forEach(function (p) {
            var detalleEtapa = (detalles[p.id] || {})[etapaId];
            if (!detalleEtapa || !Array.isArray(detalleEtapa.partidos)) return;
            detalleEtapa.partidos.forEach(function (partido) {
                var numero = Number(partido.partido);
                if (!isNaN(numero) && numero > 0) {
                    partidos[numero] = true;
                }
            });
        });

        return Object.keys(partidos).map(function (x) { return Number(x); }).sort(function (a, b) { return a - b; });
    }

    function actualizarSelectorPartido() {
        var etapaId = detalle2EtapaSel.value;
        var etapa = buscarEtapa(etapaId);
        resetSelect(detalle2PartidoSel, "Selecciona un partido");

        if (!etapaId) {
            detalle2PartidoSel.disabled = true;
            return;
        }

        if (!etapaEstaComenzada(etapa)) {
            var pendiente = document.createElement("option");
            pendiente.value = "";
            pendiente.textContent = MENSAJE_RONDA_NO_COMIENZA;
            detalle2PartidoSel.appendChild(pendiente);
            detalle2PartidoSel.disabled = true;
            return;
        }

        var partidos = partidosDisponiblesDeEtapa(etapaId);
        if (!partidos.length) {
            var vacio = document.createElement("option");
            vacio.value = "";
            vacio.textContent = "Sin partidos disponibles";
            detalle2PartidoSel.appendChild(vacio);
            detalle2PartidoSel.disabled = true;
            return;
        }

        partidos.forEach(function (numero) {
            var option = document.createElement("option");
            option.value = String(numero);
            option.textContent = etiquetaPartido(etapaId, numero);
            detalle2PartidoSel.appendChild(option);
        });
        detalle2PartidoSel.disabled = false;
    }

    function llenarSelectores() {
        participantes.forEach(function (p) {
            var optionParticipante = document.createElement("option");
            optionParticipante.value = p.id;
            optionParticipante.textContent = p.name;
            participanteSel.appendChild(optionParticipante);
        });

        etapas.forEach(function (e) {
            var optionEtapaDetalle = document.createElement("option");
            optionEtapaDetalle.value = e.id;
            optionEtapaDetalle.textContent = e.label;
            etapaSel.appendChild(optionEtapaDetalle);

            var optionEtapaPartido = document.createElement("option");
            optionEtapaPartido.value = e.id;
            optionEtapaPartido.textContent = e.label;
            detalle2EtapaSel.appendChild(optionEtapaPartido);
        });
    }

    function renderVacioDetalle() {
        content.hidden = true;
    }

    function renderVacioPartido() {
        detalle2Content.hidden = true;
    }

    function renderDetalleParticipante() {
        var participanteId = participanteSel.value;
        var etapaId = etapaSel.value;
        detalleBody.innerHTML = "";

        if (!participanteId || !etapaId) {
            renderVacioDetalle();
            return;
        }

        var participante = participantes.find(function (p) { return p.id === participanteId; });
        var etapa = buscarEtapa(etapaId);
        var detalle = (detalles[participanteId] || {})[etapaId] || null;
        var mostrarBonus = !!(etapa && etapa.show_bonus);
        renderHeaderDetalle(mostrarBonus);

        resumenParticipante.textContent = participante ? participante.name : "-";
        resumenEtapa.textContent = etapa ? etapa.label : etapaId;

        if (!etapaEstaComenzada(etapa)) {
            resumenTotal.textContent = "-";
            renderSinDataDetalle(mostrarBonus, MENSAJE_RONDA_NO_COMIENZA);
            content.hidden = false;
            return;
        }

        if (!detalle) {
            resumenTotal.textContent = "0 puntos";
            renderSinDataDetalle(mostrarBonus, "Sin pronóstico disponible para este participante en esta etapa.");
            content.hidden = false;
            return;
        }

        resumenTotal.textContent = String(detalle.total) + " puntos";
        (detalle.partidos || []).forEach(function (partido) {
            var tr = document.createElement("tr");
            appendCell(tr, etiquetaPartido(etapaId, partido.partido), "partido-cell");
            appendCell(tr, partido.pronostico || "-", "");
            appendCell(tr, partido.pauta || "-", "");
            appendCell(tr, String(partido.puntos_exactitud || 0), clasePuntos(partido.puntos_exactitud || 0));
            if (mostrarBonus) {
                appendCell(tr, String(partido.bonus || 0), clasePuntos(partido.bonus || 0));
            }
            appendCell(tr, String(partido.total || 0), clasePuntos(partido.total || 0));
            detalleBody.appendChild(tr);
        });

        content.hidden = false;
    }

    function renderDetallePartido() {
        var etapaId = detalle2EtapaSel.value;
        var etapa = buscarEtapa(etapaId);
        var mostrarBonus = !!(etapa && etapa.show_bonus);
        detalle2Body.innerHTML = "";

        if (!etapaId) {
            renderVacioPartido();
            return;
        }

        renderHeaderPartido(mostrarBonus);
        detalle2ResumenEtapa.textContent = etapa ? etapa.label : etapaId;

        if (!etapaEstaComenzada(etapa)) {
            detalle2ResumenPartido.textContent = "-";
            detalle2ResumenResultado.textContent = "-";
            renderSinDataPartido(mostrarBonus, MENSAJE_RONDA_NO_COMIENZA);
            detalle2Content.hidden = false;
            return;
        }

        if (!detalle2PartidoSel.value) {
            renderVacioPartido();
            return;
        }

        var partidoNumero = Number(detalle2PartidoSel.value);

        var filas = participantes.map(function (p) {
            var detalleEtapa = (detalles[p.id] || {})[etapaId] || null;
            var partido = null;

            if (detalleEtapa && Array.isArray(detalleEtapa.partidos)) {
                partido = detalleEtapa.partidos.find(function (x) {
                    return Number(x.partido) === partidoNumero;
                }) || null;
            }

            return {
                participante: p.name,
                pronostico: partido ? (partido.pronostico || "-") : "Sin pronóstico",
                pauta: partido ? (partido.pauta || "-") : "-",
                puntosExactitud: partido ? Number(partido.puntos_exactitud || 0) : 0,
                bonus: partido ? Number(partido.bonus || 0) : 0,
                total: partido ? Number(partido.total || 0) : 0,
            };
        });

        filas.sort(function (a, b) {
            if (b.total !== a.total) return b.total - a.total;
            return a.participante.localeCompare(b.participante, "es", { sensitivity: "base" });
        });

        var resultadoReal = filas.find(function (f) {
            return f.pauta && f.pauta !== "-" && f.pauta !== "Sin pauta";
        });

        detalle2ResumenPartido.textContent = etiquetaPartido(etapaId, partidoNumero);
        detalle2ResumenResultado.textContent = resultadoReal ? resultadoReal.pauta : "-";

        if (!filas.length) {
            renderSinDataPartido(mostrarBonus, "No hay datos disponibles para este partido.");
            detalle2Content.hidden = false;
            return;
        }

        filas.forEach(function (fila) {
            var tr = document.createElement("tr");
            appendCell(tr, fila.participante, "");
            appendCell(tr, fila.pronostico, "");
            appendCell(tr, fila.pauta, "");
            appendCell(tr, String(fila.puntosExactitud), clasePuntos(fila.puntosExactitud));
            if (mostrarBonus) {
                appendCell(tr, String(fila.bonus), clasePuntos(fila.bonus));
            }
            appendCell(tr, String(fila.total), clasePuntos(fila.total));
            detalle2Body.appendChild(tr);
        });

        detalle2Content.hidden = false;
    }

    llenarSelectores();
    participanteSel.addEventListener("change", renderDetalleParticipante);
    etapaSel.addEventListener("change", renderDetalleParticipante);
    detalle2EtapaSel.addEventListener("change", function () {
        actualizarSelectorPartido();
        renderDetallePartido();
    });
    detalle2PartidoSel.addEventListener("change", renderDetallePartido);
    renderVacioDetalle();
    renderVacioPartido();
})();
</script>
""".replace("__DETALLE_JSON__", detalle_json).replace("__RESULTADOS_JSON__", resultados_json)

    pronosticos_tabla_script = """
<script id="pronosticos-tabla-data" type="application/json">__PRONOSTICOS_TABLA_JSON__</script>
<script>
(function () {
    var dataNode = document.getElementById("pronosticos-tabla-data");
    var selector = document.getElementById("pronosticos-tabla-selector");
    var headerPartido = document.getElementById("pronostico-header-partido");
    var cells = document.querySelectorAll(".pronostico-partido-cell[data-participant-id]");
    var championChips = document.querySelectorAll(".pronostico-campeon-chip[data-country]");
    if (!dataNode || !selector || !headerPartido || !cells.length) return;

    var payload = {};
    try {
        payload = JSON.parse(dataNode.textContent || "{}");
    } catch (e) {
        console.error("No pude parsear pronosticos-tabla-data", e);
        payload = {};
    }
    var matches = Array.isArray(payload.matches) ? payload.matches : [];
    var predictions = payload.predictions || {};
    var estadosValidos = ["complete", "missing", "incomplete", "unrecognized"];

    function tiempoPartido(match) {
        var timestamp = Date.parse(String(match.datetime_chile_iso || ""));
        return Number.isFinite(timestamp) ? timestamp : null;
    }

    function seleccionarPartidoInicial(partidos, ahora) {
        var VENTANA_VISUAL_DESDE_INICIO_MS = 6 * 60 * 60 * 1000;
        var programados = partidos.map(function (match) {
            return { match: match, timestamp: tiempoPartido(match) };
        }).filter(function (item) {
            return item.timestamp !== null;
        }).sort(function (a, b) {
            return a.timestamp - b.timestamp;
        });

        var comenzadosPendientes = programados.filter(function (item) {
            return item.timestamp <= ahora && !item.match.finished;
        });
        if (comenzadosPendientes.length) {
            return comenzadosPendientes[comenzadosPendientes.length - 1].match;
        }

        var comenzadosRecientes = programados.filter(function (item) {
            return item.timestamp <= ahora &&
                   ahora < item.timestamp + VENTANA_VISUAL_DESDE_INICIO_MS;
        });
        if (comenzadosRecientes.length) {
            return comenzadosRecientes[comenzadosRecientes.length - 1].match;
        }

        var proximo = programados.find(function (item) {
            return item.timestamp > ahora;
        });
        if (proximo) return proximo.match;

        var noFinalizado = partidos.find(function (match) { return !match.finished; });
        if (noFinalizado) return noFinalizado;

        return partidos.find(function (match) { return match.id === "E06-1"; }) ||
               partidos[partidos.length - 1] || null;
    }

    function limpiar(node) {
        while (node.firstChild) node.removeChild(node.firstChild);
    }

    function pronosticoFaltante() {
        return {
            winner: "",
            mode: "",
            display: "Sin pronóstico",
            status: "missing"
        };
    }

    function renderCell(cell, pronostico) {
        var dato = pronostico || pronosticoFaltante();
        var status = estadosValidos.indexOf(dato.status) >= 0 ? dato.status : "unrecognized";
        limpiar(cell);
        estadosValidos.forEach(function (estado) {
            cell.classList.remove("pronostico-cell-" + estado);
        });
        cell.classList.add("pronostico-cell-" + status);

        var chip = document.createElement("span");
        chip.className = "pronostico-chip pronostico-chip-" + status;
        if (dato.winner && window.MundialUI &&
            typeof window.MundialUI.crearBandera === "function") {
            var bandera = window.MundialUI.crearBandera(dato.winner);
            if (bandera) chip.appendChild(bandera);
        }
        var texto = document.createElement("span");
        texto.textContent = dato.display || "Sin pronóstico";
        chip.appendChild(texto);
        cell.appendChild(chip);
    }

    function agregarBanderasCampeon() {
        if (!window.MundialUI || typeof window.MundialUI.crearBandera !== "function") return;
        championChips.forEach(function (chip) {
            var bandera = window.MundialUI.crearBandera(
                chip.getAttribute("data-country") || ""
            );
            if (bandera) chip.insertBefore(bandera, chip.firstChild);
        });
    }

    function renderMatch(match) {
        if (!match) {
            headerPartido.textContent = "Sin partido seleccionado";
            cells.forEach(function (cell) { renderCell(cell, null); });
            return;
        }
        headerPartido.textContent = match.selector_label || match.short_label || match.id;
        cells.forEach(function (cell) {
            var participantId = cell.getAttribute("data-participant-id") || "";
            var participantPredictions = predictions[participantId] || {};
            renderCell(cell, participantPredictions[match.id]);
        });
    }

    agregarBanderasCampeon();

    matches.forEach(function (match) {
        var option = document.createElement("option");
        option.value = match.id;
        option.textContent = match.selector_label;
        selector.appendChild(option);
    });

    if (!matches.length) {
        selector.disabled = true;
        renderMatch(null);
        return;
    }

    var inicial = seleccionarPartidoInicial(matches, Date.now()) || matches[0];
    selector.value = inicial.id;
    renderMatch(inicial);
    selector.addEventListener("change", function () {
        var selected = matches.find(function (match) { return match.id === selector.value; });
        renderMatch(selected || matches[0]);
    });
})();
</script>
""".replace("__PRONOSTICOS_TABLA_JSON__", pronosticos_tabla_json)

    simulacion_final_script = ""
    if mostrar_simulador_final:
        simulacion_final_script = """
<script>
(function () {
    var dataNode = document.getElementById("simulacion-final-data");
    var selector = document.getElementById("simulacion-final-selector");
    var badge = document.getElementById("simulacion-final-badge");
    var liveRegion = document.getElementById("simulacion-final-live");
    var table = document.querySelector(".tabla-posiciones.tabla-individual");
    if (!dataNode || !selector || !badge || !liveRegion || !table) return;

    var tbody = table.querySelector("tbody");
    if (!tbody) return;

    var payload = {};
    try {
        payload = JSON.parse(dataNode.textContent || "{}");
    } catch (e) {
        console.error("No pude parsear simulacion-final-data", e);
        return;
    }

    var participantes = payload.participants || {};
    var scoring = payload.scoring || {};
    var clasesPodio = ["podio-oro", "podio-plata", "podio-bronce"];
    var filas = Array.prototype.slice.call(
        tbody.querySelectorAll("tr[data-participant-id]")
    );
    var anclaFinalTbody = (
        tbody.lastChild
        && tbody.lastChild.nodeType === 3
        && !tbody.lastChild.textContent.trim()
    ) ? tbody.lastChild : null;

    function numeroConfigurado(clave, fallback) {
        var valor = Number(scoring[clave]);
        return Number.isFinite(valor) ? valor : fallback;
    }

    var puntosGanador = numeroConfigurado("winner_points", 3);
    var puntosModo = numeroConfigurado("mode_points", 1);
    var puntosBonoCampeon = numeroConfigurado("champion_bonus_points", 5);

    var originales = filas.map(function (row, indice) {
        var pid = row.getAttribute("data-participant-id") || "";
        var posCell = row.querySelector(".simulacion-pos-cell");
        var totalCell = row.querySelector(".simulacion-total-cell");
        var finalCell = row.querySelector(".simulacion-final-cell");
        var bonusCell = row.querySelector(".pronostico-campeon-puntos");
        if (!pid || !posCell || !totalCell || !finalCell || !bonusCell ||
            !participantes[pid]) return null;

        var ordenAtributo = Number(row.getAttribute("data-original-order"));
        return {
            pid: pid,
            row: row,
            posCell: posCell,
            totalCell: totalCell,
            finalCell: finalCell,
            bonusCell: bonusCell,
            classAttribute: row.getAttribute("class"),
            originalOrder: Number.isFinite(ordenAtributo) ? ordenAtributo : indice,
            posHtml: posCell.innerHTML,
            totalHtml: totalCell.innerHTML,
            finalHtml: finalCell.innerHTML,
            bonusHtml: bonusCell.innerHTML
        };
    }).filter(function (item) { return item !== null; });

    if (!originales.length) return;

    function valorVisible(valor) {
        return String(Number(valor) || 0);
    }

    function ordenarFisicamente(items, animar) {
        var posicionesIniciales = new Map();
        if (animar) {
            items.forEach(function (item) {
                posicionesIniciales.set(item.row, item.row.getBoundingClientRect().top);
            });
        }

        var fragment = document.createDocumentFragment();
        items.forEach(function (item) { fragment.appendChild(item.row); });
        tbody.insertBefore(fragment, anclaFinalTbody);

        var reducirMovimiento = (
            window.matchMedia && window.matchMedia("(prefers-reduced-motion: reduce)").matches
        );
        if (!animar || reducirMovimiento) return;

        items.forEach(function (item) {
            if (typeof item.row.animate !== "function") return;
            var posicionInicial = posicionesIniciales.get(item.row);
            var posicionFinal = item.row.getBoundingClientRect().top;
            var delta = posicionInicial - posicionFinal;
            if (!delta) return;
            item.row.animate(
                [
                    { transform: "translateY(" + delta + "px)" },
                    { transform: "translateY(0)" }
                ],
                { duration: 280, easing: "ease-out" }
            );
        });
    }

    function restaurarTabla(animar) {
        originales.forEach(function (original) {
            original.posCell.innerHTML = original.posHtml;
            original.totalCell.innerHTML = original.totalHtml;
            original.finalCell.innerHTML = original.finalHtml;
            original.bonusCell.innerHTML = original.bonusHtml;
            if (original.classAttribute === null) {
                original.row.removeAttribute("class");
            } else {
                original.row.setAttribute("class", original.classAttribute);
            }
        });

        var ordenOriginal = originales.slice().sort(function (a, b) {
            return a.originalOrder - b.originalOrder;
        });
        ordenarFisicamente(ordenOriginal, animar);
        badge.hidden = true;
        liveRegion.textContent = "Tabla actual restaurada, sin simulación";
        document.dispatchEvent(new CustomEvent("ranking-individual-actualizado"));
    }

    function compararNombre(a, b) {
        var nombreA = String(a.info.sort_name || a.info.name || "");
        var nombreB = String(b.info.sort_name || b.info.name || "");
        if (nombreA < nombreB) return -1;
        if (nombreA > nombreB) return 1;
        return a.original.originalOrder - b.original.originalOrder;
    }

    function simular(option) {
        var ganadorSimulado = option.getAttribute("data-winner") || "";
        var modoSimulado = option.getAttribute("data-mode") || "";
        var resultados = originales.map(function (original) {
            var info = participantes[original.pid] || {};
            var pronosticoFinal = info.final_prediction || {};
            var finalSimulada = 0;
            if (pronosticoFinal.winner === ganadorSimulado) {
                finalSimulada = puntosGanador;
                if (pronosticoFinal.mode && pronosticoFinal.mode === modoSimulado) {
                    finalSimulada += puntosModo;
                }
            }

            var bonoSimulado = (
                info.champion_prediction === ganadorSimulado
                ? puntosBonoCampeon
                : 0
            );
            var totalSimulado = (
                Number(info.base_without_final_and_bonus || 0)
                + finalSimulada
                + bonoSimulado
            );

            original.finalCell.textContent = valorVisible(finalSimulada);
            original.bonusCell.textContent = valorVisible(bonoSimulado);
            original.totalCell.textContent = valorVisible(totalSimulado);
            clasesPodio.forEach(function (clase) {
                original.row.classList.remove(clase);
            });

            return {
                original: original,
                info: info,
                total: totalSimulado
            };
        });

        resultados.sort(function (a, b) {
            if (b.total !== a.total) return b.total - a.total;
            return compararNombre(a, b);
        });

        var posicionAnterior = null;
        var totalAnterior = null;
        resultados.forEach(function (resultado, indice) {
            var posicion = (
                indice === 0 || resultado.total !== totalAnterior
                ? indice + 1
                : posicionAnterior
            );
            resultado.original.posCell.textContent = String(posicion);
            if (posicion === 1) resultado.original.row.classList.add("podio-oro");
            if (posicion === 2) resultado.original.row.classList.add("podio-plata");
            if (posicion === 3) resultado.original.row.classList.add("podio-bronce");
            posicionAnterior = posicion;
            totalAnterior = resultado.total;
        });

        ordenarFisicamente(
            resultados.map(function (resultado) { return resultado.original; }),
            true
        );
        badge.hidden = false;
        liveRegion.textContent = (
            "Tabla simulada: "
            + (option.getAttribute("data-winner-label") || ganadorSimulado)
            + " campeón en "
            + (option.getAttribute("data-mode-label") || modoSimulado)
        );
        document.dispatchEvent(new CustomEvent("ranking-individual-actualizado"));
    }

    selector.value = "";
    selector.addEventListener("change", function () {
        var option = selector.options[selector.selectedIndex];
        if (!option || !option.value) {
            restaurarTabla(true);
            return;
        }
        simular(option);
    });

    window.addEventListener("pageshow", function (event) {
        if (!event.persisted) return;
        selector.value = "";
        restaurarTabla(false);
    });
})();
</script>
"""

    premios_script = """
<script>
(function () {
    var dataNode = document.getElementById("premios-data");
    var openButton = document.getElementById("abrir-premios-button");
    var dialog = document.getElementById("premios-dialog");
    var dialogContent = dialog && dialog.querySelector(".premios-dialog-content");
    var poolNode = document.getElementById("premios-pozo");
    var podiumNode = document.getElementById("premios-podio");
    var roundingNote = document.getElementById("premios-redondeo-nota");
    if (!dataNode || !openButton || !dialog || !dialogContent ||
        !poolNode || !podiumNode || !roundingNote) return;

    var config = {};
    try {
        config = JSON.parse(dataNode.textContent || "{}");
    } catch (e) {
        console.error("No pude parsear premios-data", e);
        return;
    }

    var prizePool = Math.max(0, Math.round(Number(config.prize_pool) || 0));
    var prizeSlots = config.prize_slots || {"1": 60, "2": 30, "3": 10};
    var currencyFormatter = new Intl.NumberFormat("es-CL", {
        style: "currency",
        currency: "CLP",
        maximumFractionDigits: 0
    });
    var percentageFormatter = new Intl.NumberFormat("es-CL", {
        minimumFractionDigits: 0,
        maximumFractionDigits: 2
    });

    function limpiar(node) {
        while (node.firstChild) node.removeChild(node.firstChild);
    }

    function compararTexto(a, b) {
        var textoA = String(a || "").toUpperCase();
        var textoB = String(b || "").toUpperCase();
        if (textoA < textoB) return -1;
        if (textoA > textoB) return 1;
        return 0;
    }

    function leerTotal(cell) {
        if (!cell) return NaN;
        var texto = String(cell.textContent || "").trim().replace(/\\s/g, "");
        if (!texto) return NaN;
        return Number(texto.replace(/\\./g, "").replace(",", "."));
    }

    function obtenerDistribucionPremiosActual() {
        var table = document.querySelector(".tabla-posiciones.tabla-individual");
        var tbody = table && table.querySelector("tbody");
        if (!tbody) {
            return {winners: [], roundingAdjusted: false, totalAwarded: 0};
        }

        var participants = Array.prototype.slice.call(
            tbody.querySelectorAll("tr[data-participant-id]")
        ).map(function (row, originalIndex) {
            var nameCell = row.querySelector(".nombre");
            var totalCell = row.querySelector(".simulacion-total-cell");
            var total = leerTotal(totalCell);
            return {
                participantId: row.getAttribute("data-participant-id") || "",
                name: row.getAttribute("data-participant-name") ||
                      (nameCell ? nameCell.textContent.trim() : ""),
                total: total,
                originalIndex: originalIndex
            };
        }).filter(function (participant) {
            return participant.participantId && participant.name &&
                   Number.isFinite(participant.total);
        });

        participants.sort(function (a, b) {
            if (b.total !== a.total) return b.total - a.total;
            var porNombre = compararTexto(a.name, b.name);
            if (porNombre) return porNombre;
            var porId = compararTexto(a.participantId, b.participantId);
            return porId || a.originalIndex - b.originalIndex;
        });

        var winners = [];
        var index = 0;
        var groupNumber = 0;
        while (index < participants.length) {
            var end = index + 1;
            while (end < participants.length &&
                   participants[end].total === participants[index].total) {
                end += 1;
            }

            var group = participants.slice(index, end);
            var position = index + 1;
            var groupPercentage = 0;
            for (var occupied = position; occupied < position + group.length; occupied += 1) {
                groupPercentage += Number(prizeSlots[String(occupied)]) || 0;
            }

            if (groupPercentage > 0) {
                var individualPercentage = groupPercentage / group.length;
                var amountNumerator = prizePool * groupPercentage;
                var amountDenominator = 100 * group.length;
                var exactAmount = amountNumerator / amountDenominator;
                var baseAmount = Math.floor(exactAmount);
                var remainder = amountNumerator % amountDenominator;
                group.forEach(function (participant) {
                    winners.push({
                        participant_id: participant.participantId,
                        name: participant.name,
                        position: position,
                        total: participant.total,
                        percentage: individualPercentage,
                        exactAmount: exactAmount,
                        amount: baseAmount,
                        remainder: remainder,
                        denominator: amountDenominator,
                        tied: group.length > 1,
                        group: groupNumber,
                        adjustment: 0
                    });
                });
            }

            groupNumber += 1;
            index = end;
        }

        var roundedTotal = winners.reduce(function (sum, winner) {
            return sum + winner.amount;
        }, 0);
        var difference = prizePool - roundedTotal;
        var initialDifference = difference;
        if (difference && winners.length) {
            var adjustmentOrder = winners.slice().sort(function (a, b) {
                var residuosComparados = (
                    b.remainder * a.denominator -
                    a.remainder * b.denominator
                );
                if (residuosComparados) {
                    return residuosComparados;
                }
                if (a.group !== b.group) return a.group - b.group;
                return compararTexto(a.name, b.name) ||
                       compararTexto(a.participant_id, b.participant_id);
            });
            var cursor = 0;
            while (difference > 0) {
                var winner = adjustmentOrder[cursor % adjustmentOrder.length];
                winner.amount += 1;
                winner.adjustment += 1;
                difference -= 1;
                cursor += 1;
            }
        }

        winners = winners.filter(function (winner) {
            return winner.percentage > 0 && winner.amount > 0;
        });

        return {
            winners: winners,
            roundingAdjusted: initialDifference !== 0,
            adjustmentPesos: Math.abs(initialDifference),
            totalAwarded: winners.reduce(function (sum, winner) {
                return sum + winner.amount;
            }, 0)
        };
    }

    function crearTexto(clase, texto) {
        var node = document.createElement("span");
        node.className = clase;
        node.textContent = texto;
        return node;
    }

    function crearTarjetaPremio(winner) {
        var card = document.createElement("article");
        var podiumClass = winner.position >= 1 && winner.position <= 3
            ? " premios-card-pos-" + winner.position
            : "";
        card.className = "premios-card" + podiumClass;

        var placeText = winner.position + ".º lugar";
        var heading = document.createElement("div");
        heading.className = "premios-card-heading";
        heading.appendChild(crearTexto("premios-card-position", placeText));
        if (winner.tied) {
            heading.appendChild(crearTexto("premios-card-tie", "Empate"));
        }
        card.appendChild(heading);

        var name = document.createElement("h3");
        name.textContent = winner.name;
        card.appendChild(name);

        var total = document.createElement("p");
        total.className = "premios-card-total";
        total.textContent = winner.total + (winner.total === 1 ? " punto" : " puntos");
        card.appendChild(total);

        var values = document.createElement("dl");
        var percentLabel = document.createElement("dt");
        percentLabel.textContent = "Porcentaje";
        var percentValue = document.createElement("dd");
        percentValue.textContent = percentageFormatter.format(winner.percentage) + "%";
        var amountLabel = document.createElement("dt");
        amountLabel.textContent = "Premio";
        var amountValue = document.createElement("dd");
        amountValue.className = "premios-card-amount";
        amountValue.textContent = currencyFormatter.format(winner.amount);
        values.appendChild(percentLabel);
        values.appendChild(percentValue);
        values.appendChild(amountLabel);
        values.appendChild(amountValue);
        card.appendChild(values);
        return card;
    }

    function actualizarPodioPremios() {
        var distribution = obtenerDistribucionPremiosActual();
        poolNode.textContent = currencyFormatter.format(prizePool);
        limpiar(podiumNode);

        if (!distribution.winners.length) {
            var empty = document.createElement("p");
            empty.className = "premios-empty";
            empty.textContent = "No hay premios disponibles para mostrar.";
            podiumNode.appendChild(empty);
        } else {
            distribution.winners.forEach(function (winner) {
                podiumNode.appendChild(crearTarjetaPremio(winner));
            });
        }

        if (distribution.roundingAdjusted) {
            roundingNote.textContent = (
                "Se ajustó el reparto en " + distribution.adjustmentPesos +
                (distribution.adjustmentPesos === 1 ? " peso" : " pesos") +
                " para que la suma coincida exactamente con el pozo total."
            );
            roundingNote.hidden = false;
        } else {
            roundingNote.textContent = "";
            roundingNote.hidden = true;
        }
    }

    function cerrarDialogo() {
        if (typeof dialog.close === "function") {
            dialog.close();
        } else {
            dialog.removeAttribute("open");
            document.body.classList.remove("premios-modal-open");
            openButton.focus();
        }
    }

    openButton.addEventListener("click", function () {
        actualizarPodioPremios();
        document.body.classList.add("premios-modal-open");
        if (typeof dialog.showModal === "function") {
            dialog.showModal();
        } else {
            dialog.setAttribute("open", "");
        }
        var closeButton = dialog.querySelector(".premios-dialog-close-x");
        if (closeButton) closeButton.focus();
    });

    dialog.querySelectorAll(".premios-dialog-close").forEach(function (button) {
        button.addEventListener("click", cerrarDialogo);
    });

    dialog.addEventListener("click", function (event) {
        if (event.target !== dialog) return;
        var rect = dialogContent.getBoundingClientRect();
        var outside = event.clientX < rect.left || event.clientX > rect.right ||
                      event.clientY < rect.top || event.clientY > rect.bottom;
        if (outside) cerrarDialogo();
    });

    dialog.addEventListener("close", function () {
        document.body.classList.remove("premios-modal-open");
        openButton.focus();
    });

    dialog.addEventListener("cancel", function (event) {
        event.preventDefault();
        cerrarDialogo();
    });

    dialog.addEventListener("keydown", function (event) {
        if (event.key !== "Tab") return;
        var focusable = Array.prototype.slice.call(dialog.querySelectorAll(
            'button:not([disabled]), [href], input:not([disabled]), ' +
            'select:not([disabled]), textarea:not([disabled]), [tabindex]:not([tabindex="-1"])'
        )).filter(function (node) {
            return node.getClientRects().length > 0;
        });
        if (!focusable.length) return;
        var first = focusable[0];
        var last = focusable[focusable.length - 1];
        if (event.shiftKey && document.activeElement === first) {
            event.preventDefault();
            last.focus();
        } else if (!event.shiftKey && document.activeElement === last) {
            event.preventDefault();
            first.focus();
        }
    });

    document.addEventListener(
        "ranking-individual-actualizado",
        function () {
            if (dialog.open) actualizarPodioPremios();
        }
    );
})();
</script>
"""

    tendencias_script = """
<script id="tendencias-data" type="application/json">__TENDENCIAS_JSON__</script>
<script>
(function () {
    var dataNode = document.getElementById("tendencias-data");
    var selector = document.getElementById("tendencias-partido");
    var title = document.getElementById("tendencias-match-title");
    var emptyState = document.getElementById("tendencias-empty");
    var content = document.getElementById("tendencias-content");
    var teamCards = document.getElementById("tendencias-team-cards");
    var teamBar = document.getElementById("tendencias-team-bar");
    var matrix = document.getElementById("tendencias-matrix");
    if (!dataNode || !selector || !title || !emptyState || !content ||
        !teamCards || !teamBar || !matrix) return;

    var payload = {};
    try {
        payload = JSON.parse(dataNode.textContent || "{}");
    } catch (e) {
        console.error("No pude parsear tendencias-data", e);
        return;
    }
    var matches = Array.isArray(payload.matches) ? payload.matches : [];
    var modos = [
        { id: "90", label: "90'" },
        { id: "120", label: "120'" },
        { id: "PENALES", label: "Penales" }
    ];

    function limpiar(node) {
        while (node.firstChild) node.removeChild(node.firstChild);
    }

    function porcentajeSeguro(valor) {
        var numero = Number(valor);
        if (!Number.isFinite(numero) || numero < 0) return 0;
        return Math.min(100, numero);
    }

    function formatearPorcentaje(valor) {
        var redondeado = Math.round(porcentajeSeguro(valor) * 10) / 10;
        var texto = Number.isInteger(redondeado) ? String(redondeado) : redondeado.toFixed(1);
        return texto.replace(".", ",") + "%";
    }

    function agregarBandera(contenedor, nombre) {
        var ui = window.MundialUI;
        if (!ui || typeof ui.crearBandera !== "function") return;
        var bandera = ui.crearBandera(nombre);
        if (bandera) contenedor.appendChild(bandera);
    }

    function crearTarjetaEquipo(equipo, lado) {
        var card = document.createElement("div");
        card.className = "tendencias-team-card tendencias-team-" + lado.toLowerCase();

        var identity = document.createElement("div");
        identity.className = "tendencias-team-identity";
        agregarBandera(identity, equipo.name);
        var name = document.createElement("strong");
        name.textContent = equipo.name || "Por definir";
        identity.appendChild(name);

        var metric = document.createElement("div");
        metric.className = "tendencias-team-metric";
        metric.textContent = formatearPorcentaje(equipo.percentage) + " (" + String(equipo.count || 0) + ")";
        card.appendChild(identity);
        card.appendChild(metric);
        return card;
    }

    function crearSegmentoBarra(equipo, lado) {
        var segment = document.createElement("div");
        var porcentaje = porcentajeSeguro(equipo.percentage);
        segment.className = "tendencias-bar-segment tendencias-bar-" + lado.toLowerCase();
        segment.style.width = porcentaje + "%";
        segment.setAttribute(
            "aria-label",
            (equipo.name || "Equipo " + lado) + ": " + formatearPorcentaje(porcentaje) +
            " (" + String(equipo.count || 0) + ")"
        );
        if (porcentaje >= 18) {
            var label = document.createElement("span");
            label.textContent = formatearPorcentaje(porcentaje);
            segment.appendChild(label);
        }
        return segment;
    }

    function crearCeldaMatriz(equipo, modo, dato, lado) {
        var cell = document.createElement("div");
        cell.className = "tendencias-matrix-cell tendencias-matrix-" + lado.toLowerCase();
        var fill = document.createElement("div");
        fill.className = "tendencias-matrix-fill";
        fill.style.width = porcentajeSeguro(dato.percentage) + "%";
        var modeLabel = document.createElement("span");
        modeLabel.className = "tendencias-cell-mode";
        modeLabel.textContent = modo.label;
        var value = document.createElement("strong");
        value.textContent = formatearPorcentaje(dato.percentage) + " (" + String(dato.count || 0) + ")";
        cell.setAttribute(
            "aria-label",
            equipo.name + " en " + modo.label + ": " + value.textContent
        );
        cell.appendChild(fill);
        cell.appendChild(modeLabel);
        cell.appendChild(value);
        return cell;
    }

    function renderMatrix(match) {
        limpiar(matrix);
        var corner = document.createElement("div");
        corner.className = "tendencias-matrix-corner";
        matrix.appendChild(corner);
        modos.forEach(function (modo) {
            var header = document.createElement("div");
            header.className = "tendencias-matrix-header";
            header.textContent = modo.label;
            matrix.appendChild(header);
        });

        ["A", "B"].forEach(function (lado) {
            var equipo = match.teams[lado];
            var rowLabel = document.createElement("div");
            rowLabel.className = "tendencias-matrix-row-label tendencias-matrix-" + lado.toLowerCase();
            agregarBandera(rowLabel, equipo.name);
            var text = document.createElement("span");
            text.textContent = equipo.name;
            rowLabel.appendChild(text);
            matrix.appendChild(rowLabel);
            modos.forEach(function (modo) {
                var dato = match.matrix[lado][modo.id];
                matrix.appendChild(crearCeldaMatriz(equipo, modo, dato, lado));
            });
        });
    }

    function renderMatch(match) {
        if (!match) return;
        title.textContent = match.selector_label || match.short_label || "Partido";
        emptyState.hidden = true;
        content.hidden = true;

        if (!match.matchup_known) {
            emptyState.textContent = "Enfrentamiento aún no definido.";
            emptyState.hidden = false;
            return;
        }
        if (!(match.valid_team_predictions > 0) && !(match.valid_combination_predictions > 0)) {
            emptyState.textContent = "Todavía no hay pronósticos cargados para este partido.";
            emptyState.hidden = false;
            return;
        }

        limpiar(teamCards);
        limpiar(teamBar);
        teamCards.appendChild(crearTarjetaEquipo(match.teams.A, "A"));
        teamCards.appendChild(crearTarjetaEquipo(match.teams.B, "B"));
        teamBar.appendChild(crearSegmentoBarra(match.teams.A, "A"));
        teamBar.appendChild(crearSegmentoBarra(match.teams.B, "B"));
        renderMatrix(match);
        content.hidden = false;
    }

    matches.forEach(function (match) {
        var option = document.createElement("option");
        option.value = match.id;
        option.textContent = match.selector_label;
        selector.appendChild(option);
    });

    if (!matches.length) {
        title.textContent = "Sin partidos disponibles";
        emptyState.textContent = "Todavía no hay pronósticos cargados para este partido.";
        emptyState.hidden = false;
        selector.disabled = true;
        return;
    }

    var inicial = matches.find(function (match) {
        return Number(match.valid_team_predictions || 0) > 0;
    }) || matches[0];
    selector.value = inicial.id;
    renderMatch(inicial);
    selector.addEventListener("change", function () {
        var selected = matches.find(function (match) { return match.id === selector.value; });
        renderMatch(selected || matches[0]);
    });
})();
</script>
""".replace("__TENDENCIAS_JSON__", tendencias_json)

    html = f"""
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Tabla de posiciones - {titulo_competencia}</title>

<style>
:root {{
    --bg-0: #070b17;
    --bg-1: #0b1020;
    --text: #ecf2ff;
    --muted: rgba(236, 242, 255, 0.72);
    --gold-1: #f3b000;
    --gold-2: #ffdf57;
    --silver-1: #9ea8ba;
    --silver-2: #e3e9f5;
    --bronze-1: #8a4c1f;
    --bronze-2: #d08a4f;
}}

body {{
    font-family: "Trebuchet MS", "Segoe UI", "Gill Sans", sans-serif;
    background:
        radial-gradient(circle at 20% 15%, rgba(70, 122, 255, 0.20), transparent 32%),
        radial-gradient(circle at 80% 0%, rgba(255, 201, 78, 0.16), transparent 28%),
        linear-gradient(165deg, var(--bg-0), var(--bg-1), #0f1730);
    color: var(--text);
    margin: 0;
    padding: 24px;
    min-height: 100vh;
}}

.wrap {{
    max-width: 1200px;
    margin: 0 auto;
}}

.main-title {{
    margin: 0 0 8px 0;
    text-align: center;
    letter-spacing: 0.6px;
    font-size: clamp(28px, 4vw, 48px);
    color: #ffffff;
    text-shadow: 0 8px 30px rgba(0, 0, 0, 0.45);
}}

.hero {{
    margin-bottom: 14px;
}}

.sub {{
    text-align: center;
    font-size: 13px;
    color: var(--muted);
    margin-bottom: 18px;
}}

.tabla-title {{
    margin: 0 0 10px 0;
    text-align: center;
    font-size: 22px;
    letter-spacing: 0.3px;
}}

.tabla-topbar {{
    display: flex;
    justify-content: flex-end;
    margin: 0 0 12px 0;
}}

.avance-card {{
    width: min(280px, 100%);
    padding: 14px 16px;
    border-radius: 14px;
    background: rgba(12, 22, 49, 0.82);
    border: 1px solid rgba(255, 255, 255, 0.12);
    box-shadow: 0 14px 32px rgba(0, 0, 0, 0.22);
}}

.avance-label {{
    margin-bottom: 4px;
    color: var(--muted);
    font-size: 12px;
    font-weight: 700;
    letter-spacing: 0.3px;
    text-transform: uppercase;
}}

.avance-percent {{
    color: #ffffff;
    font-size: 30px;
    font-weight: 800;
    line-height: 1;
}}

.avance-progress {{
    height: 9px;
    margin-top: 12px;
    overflow: hidden;
    border-radius: 999px;
    background: rgba(0, 0, 0, 0.34);
    box-shadow: inset 0 0 0 1px rgba(255, 255, 255, 0.08);
}}

.avance-progress-fill {{
    height: 100%;
    border-radius: inherit;
    background: linear-gradient(90deg, #50d991, #ffdf57);
}}

.tabla-ranking-controls {{
    display: flex;
    align-items: end;
    justify-content: space-between;
    gap: 18px;
    margin: 0 0 20px;
}}

.tabla-ranking-controls .ranking-toggle {{
    flex: 0 1 360px;
    margin: 0;
}}

.pronosticos-tabla-control {{
    flex: 0 1 520px;
    margin-left: auto;
}}

.pronosticos-tabla-control label {{
    display: block;
    margin-bottom: 6px;
    color: var(--muted);
    font-size: 12px;
    font-weight: 800;
    letter-spacing: 0.3px;
    text-transform: uppercase;
}}

.pronosticos-tabla-selector {{
    width: 100%;
    min-height: 44px;
    box-sizing: border-box;
    padding: 10px 38px 10px 12px;
    border: 1px solid rgba(255, 255, 255, 0.16);
    border-radius: 10px;
    background: #111a35;
    color: var(--text);
    font: inherit;
}}

.tabla-acciones-individuales {{
    display: grid;
    grid-template-columns: minmax(0, 1fr) minmax(220px, 250px);
    align-items: stretch;
    gap: 16px;
    margin: 0 0 22px;
}}

.tabla-acciones-individuales[hidden] {{
    display: none;
}}

.tabla-acciones-individuales.tabla-acciones-solo-premios {{
    grid-template-columns: minmax(220px, 250px);
    justify-content: end;
}}

.premios-button {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 9px;
    min-height: 54px;
    box-sizing: border-box;
    padding: 13px 18px;
    border: 1px solid rgba(255, 236, 148, 0.72);
    border-radius: 14px;
    background: linear-gradient(135deg, #d89400, #ffdf57);
    color: #11182a;
    font: inherit;
    font-weight: 900;
    line-height: 1.25;
    cursor: pointer;
    box-shadow: 0 13px 28px rgba(151, 96, 0, 0.30);
    transition: transform 160ms ease, box-shadow 160ms ease, filter 160ms ease;
}}

.premios-button:hover {{
    filter: brightness(1.06);
    transform: translateY(-2px);
    box-shadow: 0 16px 32px rgba(151, 96, 0, 0.38);
}}

.premios-button:focus-visible {{
    outline: 3px solid #8cc8ff;
    outline-offset: 3px;
}}

.simulacion-final-control {{
    display: grid;
    grid-template-columns: minmax(0, 1fr) minmax(320px, 0.9fr);
    align-items: center;
    gap: 22px;
    margin: 0;
    padding: 20px 22px;
    border: 1px solid rgba(124, 184, 255, 0.30);
    border-radius: 16px;
    background: linear-gradient(135deg, rgba(12, 30, 67, 0.97), rgba(15, 45, 91, 0.92));
    box-shadow: 0 14px 34px rgba(0, 0, 0, 0.26);
}}

.simulacion-final-control[hidden] {{
    display: none;
}}

.simulacion-final-copy h2 {{
    margin: 0;
    color: #ffffff;
    font-size: clamp(19px, 2.4vw, 25px);
    line-height: 1.15;
}}

.simulacion-final-copy p {{
    margin: 7px 0 0;
    color: rgba(229, 239, 255, 0.78);
    font-size: 14px;
    line-height: 1.45;
}}

.simulacion-final-field {{
    min-width: 0;
}}

.simulacion-final-field label {{
    display: block;
    margin-bottom: 7px;
    color: rgba(236, 244, 255, 0.84);
    font-size: 12px;
    font-weight: 800;
    letter-spacing: 0.3px;
    text-transform: uppercase;
}}

.simulacion-final-selector-row {{
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 10px;
}}

.simulacion-final-selector {{
    flex: 1 1 320px;
    width: 100%;
    min-width: 280px;
    min-height: 46px;
    box-sizing: border-box;
    padding: 10px 38px 10px 13px;
    border: 1px solid rgba(151, 202, 255, 0.38);
    border-radius: 11px;
    background: #0b1733;
    color: #f2f7ff;
    font: inherit;
    font-weight: 700;
}}

.simulacion-final-selector:focus-visible {{
    outline: 3px solid #8cc8ff;
    outline-offset: 2px;
}}

.simulacion-final-badge {{
    display: inline-flex;
    flex: 0 0 auto;
    align-items: center;
    min-height: 28px;
    box-sizing: border-box;
    padding: 5px 9px;
    border: 1px solid rgba(255, 223, 87, 0.48);
    border-radius: 999px;
    color: #fff4bd;
    background: rgba(243, 176, 0, 0.16);
    font-size: 11px;
    font-weight: 850;
    letter-spacing: 0.25px;
    white-space: nowrap;
}}

.simulacion-final-badge[hidden] {{
    display: none;
}}

.simulacion-final-live {{
    position: absolute;
    width: 1px;
    height: 1px;
    margin: -1px;
    padding: 0;
    overflow: hidden;
    clip: rect(0 0 0 0);
    clip-path: inset(50%);
    border: 0;
    white-space: nowrap;
}}

body.premios-modal-open {{
    overflow: hidden;
}}

.premios-dialog {{
    width: min(1080px, calc(100vw - 32px));
    max-width: none;
    max-height: calc(100dvh - 32px);
    box-sizing: border-box;
    margin: auto;
    padding: 0;
    overflow: hidden;
    border: 1px solid rgba(255, 223, 87, 0.42);
    border-radius: 20px;
    background: #0b1430;
    color: var(--text);
    box-shadow: 0 28px 90px rgba(0, 0, 0, 0.72);
}}

.premios-dialog::backdrop {{
    background: rgba(2, 5, 14, 0.82);
    backdrop-filter: blur(4px);
}}

.premios-dialog-content {{
    position: relative;
    max-height: calc(100dvh - 32px);
    box-sizing: border-box;
    padding: clamp(22px, 4vw, 38px);
    overflow-y: auto;
    background:
        radial-gradient(circle at 82% 0%, rgba(255, 205, 65, 0.13), transparent 34%),
        linear-gradient(160deg, #111d40, #091127 72%);
}}

.premios-dialog-close-x {{
    position: absolute;
    z-index: 2;
    top: 14px;
    right: 14px;
    width: 42px;
    height: 42px;
    padding: 0;
    border: 1px solid rgba(255, 255, 255, 0.18);
    border-radius: 50%;
    background: rgba(3, 8, 21, 0.72);
    color: #ffffff;
    font: inherit;
    font-size: 27px;
    line-height: 1;
    cursor: pointer;
}}

.premios-dialog-close-x:hover {{
    background: rgba(255, 255, 255, 0.13);
}}

.premios-dialog-close:focus-visible {{
    outline: 3px solid #8cc8ff;
    outline-offset: 2px;
}}

.premios-dialog-header {{
    padding-right: 46px;
}}

.premios-dialog-kicker {{
    margin: 0 0 7px;
    color: #ffdf57;
    font-size: 12px;
    font-weight: 900;
    letter-spacing: 0.9px;
    text-transform: uppercase;
}}

.premios-dialog-header h2 {{
    margin: 0;
    color: #ffffff;
    font-size: clamp(25px, 4vw, 38px);
    line-height: 1.1;
}}

.premios-resumen {{
    display: grid;
    grid-template-columns: minmax(0, 1fr);
    gap: 12px;
    margin: 24px 0 28px;
}}

.premios-resumen > div {{
    min-width: 0;
    padding: 14px 16px;
    border: 1px solid rgba(255, 255, 255, 0.12);
    border-radius: 13px;
    background: rgba(5, 12, 31, 0.63);
}}

.premios-resumen-label {{
    display: block;
    margin-bottom: 5px;
    color: rgba(236, 242, 255, 0.68);
    font-size: 11px;
    font-weight: 850;
    letter-spacing: 0.45px;
    text-transform: uppercase;
}}

.premios-resumen strong {{
    display: block;
    color: #ffffff;
    font-size: 16px;
    line-height: 1.35;
    overflow-wrap: anywhere;
}}

#premios-pozo {{
    color: #ffdf57;
    font-size: 23px;
}}

.premios-podio {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
    align-items: end;
    gap: 14px;
}}

.premios-card {{
    --premios-accent: #8cc8ff;
    display: flex;
    min-width: 0;
    min-height: 170px;
    box-sizing: border-box;
    flex-direction: column;
    padding: 18px;
    border: 1px solid rgba(140, 200, 255, 0.38);
    border-color: color-mix(in srgb, var(--premios-accent) 55%, transparent);
    border-top: 7px solid var(--premios-accent);
    border-radius: 15px;
    background: rgba(12, 23, 51, 0.94);
    color: #f5f8ff;
    box-shadow: 0 15px 30px rgba(0, 0, 0, 0.27);
}}

.premios-card-pos-1 {{
    --premios-accent: #ffcf31;
    min-height: 260px;
    background: linear-gradient(155deg, rgba(103, 73, 4, 0.97), rgba(34, 29, 19, 0.98));
}}

.premios-card-pos-2 {{
    --premios-accent: #d9e1ef;
    min-height: 225px;
    background: linear-gradient(155deg, rgba(76, 88, 111, 0.97), rgba(27, 34, 51, 0.98));
}}

.premios-card-pos-3 {{
    --premios-accent: #d68a50;
    min-height: 195px;
    background: linear-gradient(155deg, rgba(105, 55, 25, 0.97), rgba(40, 25, 22, 0.98));
}}

.premios-card-heading {{
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    justify-content: space-between;
    gap: 7px;
}}

.premios-card-position {{
    color: var(--premios-accent);
    font-size: 14px;
    font-weight: 900;
    letter-spacing: 0.25px;
    text-transform: uppercase;
}}

.premios-card-tie {{
    padding: 4px 8px;
    border: 1px solid rgba(255, 255, 255, 0.28);
    border-radius: 999px;
    background: rgba(0, 0, 0, 0.24);
    color: #ffffff;
    font-size: 10px;
    font-weight: 900;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}}

.premios-card h3 {{
    margin: 15px 0 5px;
    color: #ffffff;
    font-size: clamp(18px, 2.4vw, 23px);
    line-height: 1.12;
    overflow-wrap: anywhere;
}}

.premios-card-total {{
    margin: 0 0 17px;
    color: rgba(242, 247, 255, 0.78);
    font-size: 14px;
}}

.premios-card dl {{
    display: grid;
    grid-template-columns: minmax(0, 1fr) auto;
    gap: 7px 12px;
    align-items: baseline;
    margin: auto 0 0;
}}

.premios-card dt {{
    color: rgba(236, 242, 255, 0.67);
    font-size: 12px;
}}

.premios-card dd {{
    margin: 0;
    color: #ffffff;
    font-weight: 900;
    text-align: right;
}}

.premios-card-amount {{
    color: #ffed9b !important;
    font-size: 19px;
}}

.premios-redondeo-nota {{
    margin: 22px 0 0;
    padding: 13px 15px;
    border-radius: 11px;
    color: rgba(239, 245, 255, 0.83);
    background: rgba(4, 10, 26, 0.58);
    font-size: 13px;
    line-height: 1.5;
}}

.premios-redondeo-nota {{
    margin-top: 10px;
    border: 1px solid rgba(255, 223, 87, 0.28);
    color: #fff1ad;
}}

.premios-redondeo-nota[hidden] {{
    display: none;
}}

.premios-empty {{
    grid-column: 1 / -1;
    margin: 0;
    padding: 24px;
    border: 1px dashed rgba(255, 255, 255, 0.18);
    border-radius: 13px;
    color: var(--muted);
    text-align: center;
}}

.premios-dialog-footer {{
    display: flex;
    justify-content: flex-end;
    margin-top: 22px;
}}

.premios-dialog-close-button {{
    min-width: 130px;
    min-height: 44px;
    padding: 10px 20px;
    border: 1px solid rgba(140, 200, 255, 0.42);
    border-radius: 11px;
    background: #172b58;
    color: #ffffff;
    font: inherit;
    font-weight: 850;
    cursor: pointer;
}}

.premios-dialog-close-button:hover {{
    background: #214079;
}}

.tabla-posiciones-scroll {{
    width: 100%;
    overflow-x: auto;
    overflow-y: hidden;
    border-radius: 14px;
    -webkit-overflow-scrolling: touch;
    scrollbar-color: rgba(140, 200, 255, 0.45) rgba(255, 255, 255, 0.05);
    scrollbar-width: thin;
}}

.tabla-posiciones-scroll::-webkit-scrollbar {{
    height: 8px;
}}

.tabla-posiciones-scroll::-webkit-scrollbar-track {{
    background: rgba(255, 255, 255, 0.05);
}}

.tabla-posiciones-scroll::-webkit-scrollbar-thumb {{
    border-radius: 999px;
    background: rgba(140, 200, 255, 0.42);
}}

table {{
    width: 100%;
    border-collapse: collapse;
    table-layout: fixed;
    background: #121a33;
    border-radius: 14px;
    overflow: hidden;
    box-shadow: 0 14px 36px rgba(0, 0, 0, 0.30);
}}

th, td {{
    padding: 12px;
    text-align: center;
    border-bottom: 1px solid rgba(255,255,255,0.08);
    word-wrap: break-word;
}}

tbody tr:not(:last-child) td {{
    border-bottom: 1px solid rgba(0, 0, 0, 0.28);
}}

thead tr:first-child {{
    background: #1c2753;
    font-weight: bold;
}}

thead tr:nth-child(2) {{
    background: #141c3a;
    font-size: 13px;
    opacity: 0.85;
}}

tbody tr:hover {{
    background: rgba(255,255,255,0.05);
}}

.tabla-posiciones th.total,
.tabla-posiciones td.total {{
    border: 2px solid #05070d;
}}

td.nombre {{
    text-align: center;
}}

.stage-score {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 4px;
    white-space: nowrap;
    border-radius: 999px;
    padding: 1px 5px;
    background: rgba(255, 255, 255, 0.06);
}}

.stage-medal {{
    font-size: 14px;
    line-height: 1;
}}

.stage-gold {{
    color: #ffdf57;
    font-weight: 800;
}}

.stage-silver {{
    color: #dfe7f5;
    font-weight: 800;
}}

.stage-bronze {{
    color: #e6a15a;
    font-weight: 800;
}}

tbody tr.podio-oro .stage-score,
tbody tr.podio-plata .stage-score,
tbody tr.podio-bronce .stage-score {{
    color: #111827;
    background: rgba(0, 0, 0, 0.10);
    box-shadow: inset 0 0 0 1px rgba(0, 0, 0, 0.12);
}}

.tabla-posiciones col.col-pos {{
    width: 5%;
}}

.tabla-posiciones col.col-nombre {{
    width: 22%;
}}

.tabla-posiciones col.col-total {{
    width: 9%;
}}

.tabla-posiciones col.col-puntaje {{
    width: auto;
}}

.tabla-posiciones.tabla-individual {{
    min-width: 1200px;
}}

.tabla-posiciones.tabla-familiar {{
    min-width: 760px;
}}

.tabla-posiciones.tabla-individual col.col-pos {{
    width: 4.5%;
}}

.tabla-posiciones.tabla-individual col.col-nombre {{
    width: 17%;
}}

.tabla-posiciones.tabla-individual col.col-total {{
    width: 7.5%;
}}

.tabla-posiciones.tabla-individual col.col-pronostico {{
    width: 19%;
}}

.tabla-posiciones.tabla-individual col.col-campeon {{
    width: 18%;
}}

.col-pronostico-header span,
.col-pronostico-header small {{
    display: block;
}}

.col-pronostico-header small {{
    margin-top: 4px;
    color: rgba(236, 242, 255, 0.68);
    font-size: 10px;
    font-weight: 600;
    line-height: 1.25;
    white-space: normal;
}}

.pronostico-partido-cell {{
    min-width: 165px;
    white-space: normal !important;
}}

.pronostico-campeon-cell {{
    min-width: 185px;
    white-space: normal !important;
}}

.pronostico-chip {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 6px;
    max-width: 100%;
    box-sizing: border-box;
    padding: 6px 9px;
    border: 1px solid transparent;
    border-radius: 999px;
    font-size: 12px;
    font-weight: 750;
    line-height: 1.25;
    white-space: normal;
}}

.pronostico-chip .resultados-flag-img,
.pronostico-chip .resultados-flag {{
    flex: 0 0 auto;
}}

.pronostico-campeon-nombre {{
    min-width: 0;
}}

.pronostico-campeon-separador {{
    flex: 0 0 auto;
    color: rgba(255, 255, 255, 0.72);
    font-weight: 800;
}}

.pronostico-campeon-puntos {{
    display: inline-flex;
    flex: 0 0 auto;
    align-items: center;
    justify-content: center;
    min-width: 22px;
    box-sizing: border-box;
    padding: 3px 6px;
    border: 1px solid rgba(255, 255, 255, 0.24);
    border-radius: 999px;
    color: #ffffff;
    background: rgba(3, 10, 27, 0.68);
    box-shadow: 0 1px 3px rgba(0, 0, 0, 0.30);
    font-size: 11px;
    font-weight: 850;
    line-height: 1;
}}

.pronostico-chip-complete {{
    color: #eaf3ff;
    background: rgba(49, 130, 246, 0.22);
    border-color: rgba(111, 175, 255, 0.35);
}}

.pronostico-chip-missing {{
    color: rgba(236, 242, 255, 0.66);
    background: rgba(255, 255, 255, 0.07);
    opacity: 0.82;
}}

.pronostico-chip-incomplete {{
    color: #fff2b8;
    background: rgba(225, 180, 48, 0.12);
    border-color: rgba(255, 214, 92, 0.55);
}}

.pronostico-chip-unrecognized {{
    color: #ffd7dc;
    background: rgba(205, 75, 96, 0.12);
    border-color: rgba(244, 116, 137, 0.56);
}}

.tabla-posiciones tbody tr.podio-oro .pronostico-chip,
.tabla-posiciones tbody tr.podio-plata .pronostico-chip,
.tabla-posiciones tbody tr.podio-bronce .pronostico-chip {{
    color: #ffffff;
    background: #10234d;
    border-color: rgba(255, 255, 255, 0.35);
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.30);
}}

.tabla-posiciones tbody tr.podio-oro .pronostico-chip-missing,
.tabla-posiciones tbody tr.podio-plata .pronostico-chip-missing,
.tabla-posiciones tbody tr.podio-bronce .pronostico-chip-missing {{
    color: #edf2ff;
    opacity: 0.88;
}}

.tabla-posiciones tbody tr.podio-oro .pronostico-chip-incomplete,
.tabla-posiciones tbody tr.podio-plata .pronostico-chip-incomplete,
.tabla-posiciones tbody tr.podio-bronce .pronostico-chip-incomplete {{
    color: #ffffff;
    border-color: rgba(255, 215, 92, 0.88);
}}

.tabla-posiciones tbody tr.podio-oro .pronostico-chip-unrecognized,
.tabla-posiciones tbody tr.podio-plata .pronostico-chip-unrecognized,
.tabla-posiciones tbody tr.podio-bronce .pronostico-chip-unrecognized {{
    color: #ffffff;
    border-color: rgba(255, 126, 145, 0.88);
}}

.tabla-posiciones th,
.tabla-posiciones td {{
    overflow-wrap: normal;
    word-wrap: normal;
}}

.tabla-posiciones th:nth-child(2),
.tabla-posiciones td.nombre {{
    white-space: normal;
    word-break: keep-all;
    overflow-wrap: normal;
    word-wrap: normal;
    hyphens: none;
    line-height: 1.25;
}}

.tabla-posiciones thead tr:nth-child(2) th,
.tabla-posiciones tbody td:not(.nombre) {{
    white-space: nowrap;
    word-break: normal;
    overflow-wrap: normal;
    word-wrap: normal;
}}

/* Estilo podio dentro de la tabla existente */
tbody tr.podio-oro {{
    background: linear-gradient(90deg, #ffd700, #ffea00, #ffd700);
    color: #1a1a1a;
    font-weight: 700;
}}

tbody tr.podio-plata {{
    background: linear-gradient(90deg, #cfcfcf, #f5f5f5, #cfcfcf);
    color: #1a1a1a;
    font-weight: 700;
}}

tbody tr.podio-bronce {{
    background: linear-gradient(90deg, #cd7f32, #e6a15a, #cd7f32);
    color: #1a1a1a;
    font-weight: 700;
}}

tbody tr.podio-oro:hover,
tbody tr.podio-plata:hover,
tbody tr.podio-bronce:hover {{
    filter: brightness(1.05);
}}

.tendencias-wrap {{
    margin-top: 34px;
    padding: 22px;
    border-radius: 16px;
    background: rgba(12, 22, 49, 0.88);
    border: 1px solid rgba(255, 255, 255, 0.09);
    box-shadow: 0 14px 36px rgba(0, 0, 0, 0.25);
}}

.tendencias-sub {{
    margin: -8px 0 18px;
    color: var(--muted);
}}

.tendencias-selector-field {{
    width: min(100%, 640px);
    margin: 0 auto 18px;
}}

.tendencias-selector-field label {{
    display: block;
    margin-bottom: 7px;
    color: var(--muted);
    font-size: 13px;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: 0.35px;
}}

.tendencias-select {{
    width: 100%;
    min-height: 44px;
    box-sizing: border-box;
    padding: 10px 38px 10px 12px;
    border: 1px solid rgba(255, 255, 255, 0.16);
    border-radius: 10px;
    background: #111a35;
    color: var(--text);
    font: inherit;
}}

.tendencias-panel {{
    padding: 20px;
    border-radius: 14px;
    background: rgba(5, 10, 24, 0.48);
    border: 1px solid rgba(255, 255, 255, 0.07);
}}

.tendencias-match-title {{
    margin: 0 0 18px;
    text-align: center;
    font-size: clamp(18px, 3vw, 24px);
}}

.tendencias-empty {{
    padding: 28px 18px;
    border: 1px dashed rgba(255, 255, 255, 0.18);
    border-radius: 12px;
    color: var(--muted);
    text-align: center;
}}

.tendencias-chart-title {{
    margin: 0 0 12px;
    color: #ffffff;
    font-size: 15px;
    font-weight: 800;
}}

.tendencias-team-cards {{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 12px;
    margin-bottom: 12px;
}}

.tendencias-team-card {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
    min-width: 0;
    padding: 13px 15px;
    border-radius: 12px;
    border: 1px solid rgba(255, 255, 255, 0.10);
}}

.tendencias-team-a {{
    background: rgba(49, 130, 246, 0.16);
}}

.tendencias-team-b {{
    background: rgba(239, 101, 128, 0.16);
}}

.tendencias-team-identity {{
    display: flex;
    align-items: center;
    gap: 8px;
    min-width: 0;
}}

.tendencias-team-identity strong {{
    overflow: hidden;
    text-overflow: ellipsis;
}}

.tendencias-team-metric {{
    flex: 0 0 auto;
    font-weight: 900;
}}

.tendencias-team-bar {{
    display: flex;
    width: 100%;
    min-height: 48px;
    overflow: hidden;
    border-radius: 12px;
    background: rgba(255, 255, 255, 0.06);
    box-shadow: inset 0 0 0 1px rgba(255, 255, 255, 0.09);
}}

.tendencias-bar-segment {{
    display: flex;
    align-items: center;
    justify-content: center;
    overflow: hidden;
    transition: width 180ms ease;
}}

.tendencias-bar-segment span {{
    padding: 0 8px;
    color: #ffffff;
    font-weight: 900;
    white-space: nowrap;
    text-shadow: 0 1px 3px rgba(0, 0, 0, 0.45);
}}

.tendencias-bar-a {{
    background: linear-gradient(90deg, #2563eb, #4f9cff);
}}

.tendencias-bar-b {{
    background: linear-gradient(90deg, #e64f73, #f47f98);
}}

.tendencias-matrix-title {{
    margin-top: 26px;
}}

.tendencias-matrix-scroll {{
    overflow-x: auto;
    padding-bottom: 3px;
}}

.tendencias-matrix {{
    display: grid;
    grid-template-columns: minmax(120px, 1.25fr) repeat(3, minmax(110px, 1fr));
    gap: 8px;
    min-width: 560px;
}}

.tendencias-matrix-corner,
.tendencias-matrix-header {{
    padding: 6px 8px;
    color: var(--muted);
    font-size: 13px;
    font-weight: 800;
    text-align: center;
}}

.tendencias-matrix-row-label {{
    display: flex;
    align-items: center;
    gap: 7px;
    min-width: 0;
    padding: 12px;
    border-radius: 10px;
    font-weight: 800;
}}

.tendencias-matrix-row-label span {{
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}}

.tendencias-matrix-row-label.tendencias-matrix-a {{
    background: rgba(49, 130, 246, 0.16);
}}

.tendencias-matrix-row-label.tendencias-matrix-b {{
    background: rgba(239, 101, 128, 0.16);
}}

.tendencias-matrix-cell {{
    position: relative;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    min-height: 62px;
    overflow: hidden;
    border-radius: 10px;
    background: rgba(255, 255, 255, 0.045);
    border: 1px solid rgba(255, 255, 255, 0.08);
}}

.tendencias-matrix-fill {{
    position: absolute;
    inset: 0 auto 0 0;
    opacity: 0.27;
}}

.tendencias-matrix-a .tendencias-matrix-fill {{
    background: #4f9cff;
}}

.tendencias-matrix-b .tendencias-matrix-fill {{
    background: #f47f98;
}}

.tendencias-cell-mode,
.tendencias-matrix-cell strong {{
    position: relative;
    z-index: 1;
}}

.tendencias-cell-mode {{
    color: var(--muted);
    font-size: 11px;
    font-weight: 700;
}}

.tendencias-matrix-cell strong {{
    margin-top: 3px;
    font-size: 14px;
}}

@media (max-width: 600px) {{
    .tendencias-wrap {{
        padding: 16px;
    }}
    .tendencias-panel {{
        padding: 14px;
    }}
    .tendencias-team-cards {{
        grid-template-columns: 1fr;
    }}
    .tendencias-team-card {{
        padding: 12px;
    }}
}}

.resultados-wrap {{
    margin-top: 34px;
    padding: 22px;
    border-radius: 16px;
    background: rgba(12, 22, 49, 0.82);
    border: 1px solid rgba(255, 255, 255, 0.08);
    box-shadow: 0 14px 36px rgba(0, 0, 0, 0.25);
}}

.resultados-sub {{
    margin: -8px 0 18px 0;
    color: var(--muted);
    font-size: 14px;
}}

.resultados-layout {{
    display: grid;
    grid-template-columns: minmax(220px, 280px) minmax(0, 1fr);
    gap: 18px;
    align-items: start;
}}

.resultados-panel-title {{
    display: flex;
    justify-content: space-between;
    gap: 12px;
    align-items: baseline;
    margin-bottom: 10px;
}}

.resultados-panel-title strong {{
    font-size: 16px;
}}

.resultados-count {{
    color: var(--muted);
    font-size: 13px;
    white-space: nowrap;
}}

.resultados-note,
.resultados-empty {{
    border: 1px solid rgba(255, 255, 255, 0.09);
    border-radius: 12px;
    padding: 12px 14px;
    background: rgba(255, 255, 255, 0.06);
    color: var(--muted);
    font-size: 14px;
}}

.resultados-note {{
    margin-bottom: 10px;
}}

.resultados-table th,
.resultados-table td {{
    vertical-align: middle;
}}

.resultados-match-cell strong,
.resultados-match-cell span {{
    display: block;
}}

.resultados-match-cell span {{
    margin-top: 3px;
    color: var(--muted);
    font-size: 12px;
}}

.resultados-match-cell .resultados-schedule-line {{
    margin-top: 4px;
    color: var(--muted);
    font-size: 12px;
    line-height: 1.25;
}}

.resultados-match-cell .resultados-venue-line {{
    margin-top: 2px;
    color: rgba(236, 242, 255, 0.50);
    font-size: 11px;
    line-height: 1.25;
}}

.resultados-team {{
    display: inline-flex;
    align-items: center;
    gap: 8px;
    line-height: 1.25;
    font-weight: 600;
}}

.resultados-flag {{
    min-width: 1.45em;
    font-size: 18px;
    text-align: center;
}}

.resultados-flag-img {{
    width: 24px;
    height: 18px;
    flex: 0 0 24px;
    border-radius: 3px;
    object-fit: cover;
    box-shadow: 0 0 0 1px rgba(255, 255, 255, 0.18);
}}

.resultados-team-win {{
    color: #91f5a8;
    font-weight: 800;
}}

.resultados-team-pending {{
    color: rgba(236, 242, 255, 0.55);
    font-weight: 500;
}}

.resultados-result-cell {{
    min-width: 170px;
}}

.resultados-badge {{
    display: inline-flex;
    align-items: center;
    border-radius: 999px;
    padding: 4px 9px;
    margin-bottom: 6px;
    font-size: 12px;
    font-weight: 800;
    letter-spacing: 0.2px;
}}

.resultados-badge-winner {{
    background: rgba(126, 242, 158, 0.14);
    color: #91f5a8;
    border: 1px solid rgba(126, 242, 158, 0.26);
}}

.resultados-badge-draw,
.resultados-badge-text {{
    background: rgba(255, 223, 87, 0.13);
    color: #ffdf57;
    border: 1px solid rgba(255, 223, 87, 0.25);
}}

.resultados-badge-pending {{
    background: rgba(215, 222, 239, 0.10);
    color: rgba(236, 242, 255, 0.62);
    border: 1px solid rgba(215, 222, 239, 0.14);
}}

.resultados-result-line {{
    display: flex;
    align-items: center;
    gap: 4px;
    flex-wrap: wrap;
}}

.resultados-mode {{
    color: var(--muted);
}}

.resultados-pendiente {{
    opacity: 0.76;
}}

.detalle-wrap {{
    margin-top: 34px;
    padding: 22px;
    border-radius: 16px;
    background: rgba(12, 22, 49, 0.82);
    border: 1px solid rgba(255, 255, 255, 0.08);
    box-shadow: 0 14px 36px rgba(0, 0, 0, 0.25);
}}

.detalle-wrap-sec {{
    margin-top: 18px;
}}

.detalle-title {{
    margin: 0 0 16px 0;
    text-align: left;
    font-size: clamp(20px, 2.4vw, 28px);
}}

.detalle-controls {{
    display: grid;
    grid-template-columns: repeat(2, minmax(220px, 1fr));
    gap: 14px;
}}

.detalle-field {{
    display: flex;
    flex-direction: column;
    gap: 6px;
}}

.detalle-field label {{
    font-size: 13px;
    color: var(--muted);
}}

.detalle-select {{
    width: 100%;
    border: 1px solid rgba(255, 255, 255, 0.16);
    border-radius: 10px;
    padding: 10px 12px;
    background: #0e1a3a;
    color: var(--text);
    font-size: 15px;
}}

.detalle-select:disabled {{
    opacity: 0.66;
    cursor: not-allowed;
}}

.detalle-summary {{
    margin-top: 16px;
    display: grid;
    grid-template-columns: repeat(3, minmax(170px, 1fr));
    gap: 10px;
}}

.detalle-card {{
    background: rgba(255, 255, 255, 0.07);
    border: 1px solid rgba(255, 255, 255, 0.08);
    border-radius: 12px;
    padding: 12px 14px;
}}

.detalle-label {{
    font-size: 12px;
    color: var(--muted);
    margin-bottom: 4px;
}}

.detalle-value {{
    font-size: 17px;
    font-weight: 700;
}}

.detalle-total {{
    color: #ffdf57;
}}

.detalle-table-wrap {{
    margin-top: 14px;
    overflow-x: auto;
}}

.detalle-table {{
    width: 100%;
    border-collapse: collapse;
    background: #101a39;
    border-radius: 12px;
    overflow: hidden;
}}

.detalle-table th,
.detalle-table td {{
    padding: 10px 8px;
    border-bottom: 1px solid rgba(255, 255, 255, 0.08);
    text-align: left;
    font-size: 14px;
    vertical-align: top;
}}

.detalle-table th {{
    background: #1a2654;
    color: #f6f8ff;
    font-weight: 700;
}}

.detalle-table td.num {{
    text-align: center;
    white-space: nowrap;
}}

.detalle-table td.partido-cell {{
    min-width: 180px;
    color: #dfe7f5;
    font-weight: 700;
}}

.detalle-table th.num {{
    text-align: center;
    white-space: normal;
    line-height: 1.2;
    word-break: break-word;
}}

.pts-pos {{
    color: #7ef29e;
    font-weight: 700;
}}

.pts-zero {{
    color: #d7deef;
}}

@media (max-width: 900px) {{
    .tabla-acciones-individuales,
    .tabla-acciones-individuales.tabla-acciones-solo-premios {{
        grid-template-columns: 1fr;
        justify-content: stretch;
    }}

    .premios-button {{
        width: 100%;
    }}

    .simulacion-final-control {{
        grid-template-columns: 1fr;
        gap: 16px;
    }}

    table {{
        font-size: 13px;
    }}

    th, td {{
        padding: 10px 8px;
    }}

    .detalle-controls {{
        grid-template-columns: 1fr;
    }}

    .resultados-layout {{
        grid-template-columns: 1fr;
    }}

    .resultados-panel-title {{
        align-items: flex-start;
        flex-direction: column;
    }}

    .detalle-summary {{
        grid-template-columns: 1fr;
    }}

    .detalle-table th,
    .detalle-table td {{
        font-size: 13px;
    }}

    .tabla-topbar {{
        justify-content: center;
    }}
}}

@media (max-width: 600px) {{
    .tabla-topbar {{
        justify-content: stretch;
    }}

    .tabla-ranking-controls {{
        flex-direction: column;
        align-items: stretch;
        gap: 12px;
    }}

    .tabla-ranking-controls .ranking-toggle,
    .pronosticos-tabla-control {{
        flex-basis: auto;
        width: 100%;
        margin-left: 0;
    }}

    .simulacion-final-control {{
        grid-template-columns: 1fr;
        gap: 16px;
        padding: 18px 16px;
    }}

    .simulacion-final-selector-row {{
        align-items: stretch;
        flex-direction: column;
    }}

    .simulacion-final-selector {{
        flex-basis: auto;
        min-width: 0;
    }}

    .simulacion-final-badge {{
        align-self: flex-start;
    }}

    .premios-dialog {{
        width: calc(100vw - 16px);
        max-height: calc(100dvh - 16px);
        border-radius: 15px;
    }}

    .premios-dialog-content {{
        max-height: calc(100dvh - 16px);
        padding: 22px 15px 18px;
    }}

    .premios-dialog-header {{
        padding-right: 43px;
    }}

    .premios-resumen {{
        grid-template-columns: 1fr;
        margin-top: 20px;
    }}

    .premios-podio {{
        grid-template-columns: 1fr;
    }}

    .premios-card-pos-1 {{
        min-height: 230px;
    }}

    .premios-card-pos-2 {{
        min-height: 205px;
    }}

    .premios-card-pos-3 {{
        min-height: 180px;
    }}

    .premios-dialog-close-button {{
        width: 100%;
    }}

    .avance-card {{
        width: 100%;
        box-sizing: border-box;
    }}

    .tabla-posiciones {{
        font-size: 12px;
    }}

    .tabla-posiciones col.col-nombre {{
        width: 23%;
    }}

    .tabla-posiciones col.col-total {{
        width: 10%;
    }}

    .tabla-posiciones th,
    .tabla-posiciones td {{
        padding: 8px 4px;
    }}

    .tabla-posiciones th:first-child,
    .tabla-posiciones td:first-child,
    .tabla-posiciones th:nth-child(n+3),
    .tabla-posiciones td:nth-child(n+3) {{
        padding-left: 3px;
        padding-right: 3px;
    }}

    .tabla-posiciones th:nth-child(2),
    .tabla-posiciones td.nombre {{
        font-size: 11px;
        padding-left: 5px;
        padding-right: 5px;
    }}

    .stage-score {{
        gap: 2px;
        padding-left: 3px;
        padding-right: 3px;
    }}

    .stage-medal {{
        font-size: 12px;
    }}
}}

{ranking_familiar_css}

</style>
</head>

<body>
<div class="wrap">
<div class="hero">
<h1 class="main-title">{titulo_competencia}</h1>
<div class="sub">Generado: {now}</div>
</div>

<div class="tabla-topbar">
    <div class="avance-card" aria-label="Avance del Mundial">
        <div class="avance-label">Avance del Mundial (% de puntos repartidos)</div>
        <div class="avance-percent">{porcentaje_display}</div>
        <div class="avance-progress" aria-hidden="true">
            <div class="avance-progress-fill" style="width: {progreso_width:.1f}%;"></div>
        </div>
    </div>
</div>

<div class="tabla-ranking-controls">
{ranking_toggle_html}
{pronosticos_control_html}
</div>
{tabla_acciones_individuales_html}
{rankings_html}
{premios_dialog_html}

<section class="tendencias-wrap" id="tendencias-section">
<h2 class="detalle-title">Tendencia de pronósticos</h2>
<p class="tendencias-sub">Distribución de los pronósticos de los participantes desde cuartos de final</p>

<div class="tendencias-selector-field">
    <label for="tendencias-partido">Partido</label>
    <select id="tendencias-partido" class="tendencias-select"></select>
</div>

<div class="tendencias-panel">
    <h3 id="tendencias-match-title" class="tendencias-match-title">Partido</h3>
    <div id="tendencias-empty" class="tendencias-empty" hidden></div>
    <div id="tendencias-content" hidden>
        <h4 class="tendencias-chart-title">Qué equipo avanza</h4>
        <div id="tendencias-team-cards" class="tendencias-team-cards"></div>
        <div id="tendencias-team-bar" class="tendencias-team-bar"
             role="img" aria-label="Distribución de pronósticos por equipo"></div>

        <h4 class="tendencias-chart-title tendencias-matrix-title">Cómo avanza cada equipo</h4>
        <div class="tendencias-matrix-scroll">
            <div id="tendencias-matrix" class="tendencias-matrix"
                 role="grid" aria-label="Distribución por equipo y modo"></div>
        </div>
    </div>
</div>
</section>

<section class="resultados-wrap" id="resultados-section">
<h2 class="detalle-title">Resultados actualizados del Mundial</h2>
<p class="resultados-sub">Resultados oficiales con última actualización en {now}</p>

<div class="resultados-layout">
    <div class="detalle-field">
        <label for="resultados-etapa">Etapa</label>
        <select id="resultados-etapa" class="detalle-select">
            <option value="">Selecciona una etapa</option>
        </select>
    </div>

    <div class="resultados-board">
        <div class="resultados-panel-title">
            <strong id="resultados-stage-label">Etapa</strong>
            <span class="resultados-count" id="resultados-stage-count">0 partidos</span>
        </div>
        <div id="resultados-note" class="resultados-note" hidden></div>
        <div id="resultados-empty" class="resultados-empty" hidden></div>
        <div class="detalle-table-wrap" id="resultados-table-wrap" hidden>
            <table class="detalle-table resultados-table">
                <thead>
                    <tr>
                        <th>Partido / horario</th>
                        <th>Equipo A</th>
                        <th>Equipo B</th>
                        <th>Resultado oficial</th>
                    </tr>
                </thead>
                <tbody id="resultados-body"></tbody>
            </table>
        </div>
    </div>
</div>
</section>

<section class="detalle-wrap" id="detalle-section">
<h2 class="detalle-title">Detalle por participante y etapa</h2>

<div class="detalle-controls">
    <div class="detalle-field">
        <label for="detalle-participante">Participante</label>
        <select id="detalle-participante" class="detalle-select">
            <option value="">Selecciona un participante</option>
        </select>
    </div>
    <div class="detalle-field">
        <label for="detalle-etapa">Etapa</label>
        <select id="detalle-etapa" class="detalle-select">
            <option value="">Selecciona una etapa</option>
        </select>
    </div>
</div>

<div id="detalle-content" hidden>
    <div class="detalle-summary">
        <div class="detalle-card">
            <div class="detalle-label">Participante</div>
            <div class="detalle-value" id="detalle-resumen-participante">-</div>
        </div>
        <div class="detalle-card">
            <div class="detalle-label">Etapa</div>
            <div class="detalle-value" id="detalle-resumen-etapa">-</div>
        </div>
        <div class="detalle-card">
            <div class="detalle-label">Total etapa</div>
            <div class="detalle-value detalle-total" id="detalle-resumen-total">0 puntos</div>
        </div>
    </div>

    <div class="detalle-table-wrap">
        <table class="detalle-table">
            <thead>
                <tr id="detalle-head-row"></tr>
            </thead>
            <tbody id="detalle-body"></tbody>
        </table>
    </div>
</div>
</section>

<section class="detalle-wrap detalle-wrap-sec" id="detalle2-section">
<h2 class="detalle-title">Detalle por etapa y partido</h2>

<div class="detalle-controls">
    <div class="detalle-field">
        <label for="detalle2-etapa">Etapa</label>
        <select id="detalle2-etapa" class="detalle-select">
            <option value="">Selecciona una etapa</option>
        </select>
    </div>
    <div class="detalle-field">
        <label for="detalle2-partido">Partido</label>
        <select id="detalle2-partido" class="detalle-select" disabled>
            <option value="">Selecciona un partido</option>
        </select>
    </div>
</div>

<div id="detalle2-content" hidden>
    <div class="detalle-summary">
        <div class="detalle-card">
            <div class="detalle-label">Etapa</div>
            <div class="detalle-value" id="detalle2-resumen-etapa">-</div>
        </div>
        <div class="detalle-card">
            <div class="detalle-label">Partido</div>
            <div class="detalle-value" id="detalle2-resumen-partido">-</div>
        </div>
        <div class="detalle-card">
            <div class="detalle-label">Resultado Real</div>
            <div class="detalle-value" id="detalle2-resumen-resultado">-</div>
        </div>
    </div>

    <div class="detalle-table-wrap">
        <table class="detalle-table">
            <thead>
                <tr id="detalle2-head-row"></tr>
            </thead>
            <tbody id="detalle2-body"></tbody>
        </table>
    </div>
</div>
</section>

{ranking_toggle_script}
{detalle_script}
{pronosticos_tabla_script}
{simulacion_final_data_html}
{simulacion_final_script}
{premios_data_html}
{premios_script}
{tendencias_script}
</div>
</body>
</html>
"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

def render_portada_html(out_path):
    html = """<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Pollas Mundialeras</title>
<style>
body {
    font-family: "Trebuchet MS", "Segoe UI", sans-serif;
    margin: 0;
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    background: linear-gradient(135deg, #0a1124, #172a59);
    color: #eef3ff;
}
.wrap {
    width: min(560px, calc(100% - 40px));
    background: rgba(8, 16, 37, 0.8);
    border: 1px solid rgba(255, 255, 255, 0.12);
    border-radius: 14px;
    padding: 26px;
    box-shadow: 0 14px 34px rgba(0, 0, 0, 0.35);
}
h1 {
    margin: 0 0 8px;
    text-align: center;
}
p {
    margin: 0 0 18px;
    text-align: center;
    color: rgba(238, 243, 255, 0.78);
}
.links {
    display: grid;
    gap: 10px;
}
a {
    display: block;
    text-decoration: none;
    text-align: center;
    padding: 12px;
    border-radius: 10px;
    background: #243f86;
    color: #ffffff;
    font-weight: 700;
}
a:hover {
    background: #2f50a6;
}
</style>
</head>
<body>
<main class="wrap">
<h1>Pollas Mundialeras</h1>
<p>Selecciona una competencia para ver la tabla.</p>
<div class="links">
<a href="./familia/index.html">Ver polla familia</a>
<a href="./curso/index.html">Ver polla curso</a>
</div>
</main>
</body>
</html>
"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def generar_competencia(nombre_competencia, nombre_carpeta_participantes,
                        subcarpeta_salida, calendario_por_etapa=None,
                        usar_ranking_familiar=False, pozo_premios=0):
    nombre_competencia = str(nombre_competencia).strip()
    pozo_premios = max(0, int(pozo_premios or 0))
    titulo_competencia = (
        nombre_competencia
        if nombre_competencia.lower().startswith("polla ")
        else f"Polla {nombre_competencia}"
    )

    print(f"\nGenerando competencia: {nombre_competencia}")
    etiqueta_log_premios = (
        "Familia" if nombre_competencia.lower() == "familia" else nombre_competencia
    )
    print(
        f"[{etiqueta_log_premios}] Pozo de premios: "
        f"{formatear_monto_clp_python(pozo_premios)}"
    )
    datos = {}

    # 1) Cargar pautas oficiales desde carpeta Pauta (compartida)
    try:
        carpeta_pauta = resolver_carpeta_pauta(CARPETA)
        (
            pautas_por_etapa,
            fuentes_pauta,
            etapas_faltantes,
            archivos_ignorados,
            enfrentamientos_por_etapa,
            campeon_real_pauta,
            enfrentamientos_detalle_por_etapa,
        ) = cargar_pautas_desde_excel(carpeta_pauta)
    except Exception as e:
        print(f"[{nombre_competencia}] ERROR cargando pauta compartida: {e}")
        return False

    print(f"[{nombre_competencia}] Pautas cargadas desde: {carpeta_pauta}")
    for etapa in sorted(fuentes_pauta.keys(), key=clave_orden_etapa):
        print(f"  - {etapa}: {fuentes_pauta[etapa]}")

    if archivos_ignorados:
        print(f"\n[{nombre_competencia}] Aviso: se ignoraron archivos en pauta:")
        for fn in archivos_ignorados:
            print(f"  - {fn}")

    if etapas_faltantes:
        print(f"\n[{nombre_competencia}] Aviso: faltan pautas para estas etapas:")
        print("  " + ", ".join(etapas_faltantes))
        print("  Se asignará 0 en esas etapas y quedará registrado como error por participante.")

    # 2) Cargar pronósticos de la carpeta específica de la competencia
    try:
        carpeta_participantes = resolver_carpeta_por_nombre(
            CARPETA, nombre_carpeta_participantes
        )
    except Exception as e:
        print(f"[{nombre_competencia}] ERROR cargando carpeta de participantes: {e}")
        return False

    registros, avisos_pronostico, etapas_vacias = cargar_archivos_pronostico(carpeta_participantes)

    if avisos_pronostico:
        print(f"\n[{nombre_competencia}] Aviso: se ignoraron archivos de pronóstico:")
        for aviso in avisos_pronostico:
            print(f"  - {aviso}")

    if etapas_vacias:
        print(f"\n[{nombre_competencia}] Aviso: hay carpetas de etapa sin archivos de pronóstico:")
        print("  " + ", ".join(etapas_vacias))

    if not registros:
        print(
            f"[{nombre_competencia}] ERROR: no encontré pronósticos válidos en "
            f"'{nombre_carpeta_participantes}'. Usa carpetas por etapa (ej: E01 o etapa 01) "
            "y archivos como Nombre.xlsx."
        )
        return False

    procesados = set()

    for reg in registros:
        etapa = reg["etapa"]
        nombre = reg["nombre"]
        pid = reg["pid"]
        ruta = reg["ruta"]
        fn = reg["archivo"]

        datos.setdefault(pid, {
            "nombre": nombre,
            "scores": {e: 0 for e in ETAPAS.keys()},
            "detalle_etapas": {},
            "pronosticos_elim": {},
            "campeon_pred": None,
            "errores": []
        })

        datos[pid]["nombre"] = nombre

        clave_archivo = (pid, etapa)
        if clave_archivo in procesados:
            datos[pid]["errores"].append(
                f"{fn}: archivo duplicado para {nombre} en etapa {etapa} (se ignoró)."
            )
            continue
        procesados.add(clave_archivo)

        if etapa in {"E04", "E05", "E06"}:
            try:
                datos[pid]["pronosticos_elim"][etapa] = (
                    leer_pronosticos_eliminatoria_crudos(ruta, etapa)
                )
            except Exception as e:
                datos[pid]["errores"].append(
                    f"{fn}: error leyendo pronóstico eliminatorio: {e}"
                )

        try:
            detalle_etapa = calcular_detalle_etapa(ruta, etapa, pautas_por_etapa)
            pts_etapa = detalle_etapa["total_etapa"]
            datos[pid]["detalle_etapas"][etapa] = detalle_etapa
        except Exception as e:
            pts_etapa = 0
            datos[pid]["errores"].append(f"{fn}: {e}")

        datos[pid]["scores"][etapa] = pts_etapa

        # Guardar campeón pronosticado SOLO desde E01 (B4)
        if etapa == "E01":
            try:
                datos[pid]["campeon_pred"] = leer_campeon_predicho_desde_e01(ruta)
            except Exception as e:
                datos[pid]["campeon_pred"] = None
                datos[pid]["errores"].append(f"{fn}: error leyendo campeón en B4: {e}")

    partidos_clave = construir_partidos_clave_eliminatorias(
        enfrentamientos_detalle_por_etapa=enfrentamientos_detalle_por_etapa,
        pautas_por_etapa=pautas_por_etapa,
        calendario_por_etapa=calendario_por_etapa,
    )
    tendencias_payload = construir_tendencias_eliminatorias(
        datos=datos,
        partidos_clave=partidos_clave,
    )
    pronosticos_tabla_payload = construir_pronosticos_tabla_payload(
        datos=datos,
        partidos_clave=partidos_clave,
    )

    # 3) Orden de etapas (ya existe para toda la impresión)
    etapas_ordenadas = sorted(ETAPAS.keys(), key=clave_orden_etapa)

    # 4) Parámetros de impresión "tipo Excel"
    W_POS = 4
    W_NOMBRE = 12
    W_ETAPA = 12
    W_BONUS = 15
    W_TOTAL = 10

    max_por_etapa = {}
    for e, cfg in ETAPAS.items():
        if cfg["tipo"] == "GRUPOS":
            max_por_etapa[e] = cfg["n_partidos"] * cfg["ppp"]
        else:
            max_por_etapa[e] = cfg["n_partidos"] * (cfg["ppp"] + 1)

    max_bonus = BONUS_PTS
    max_total = sum(max_por_etapa.values()) + max_bonus

    def c(text, w):
        return f"{str(text):^{w}}"

    def l(text, w):
        return f"{str(text):<{w}}"

    ancho_tabla = (
        W_POS + 1 + W_NOMBRE + 1 +
        len(etapas_ordenadas) * (W_ETAPA + 1) +
        (W_BONUS + 1) + W_TOTAL
    )

    # Normalización simple para comparar campeones
    def norm(x):
        if x is None:
            return ""
        return str(x).strip().upper()

    campeon_real_oficial = campeon_real_pauta
    if not norm(campeon_real_oficial):
        campeon_real_oficial = CAMPEON_REAL_MANUAL
    campeon_real_norm = norm(campeon_real_oficial)

    if norm(campeon_real_pauta):
        print(f"\n[{nombre_competencia}] Campeón oficial desde pauta E01 (B4): {campeon_real_pauta}")
    elif norm(CAMPEON_REAL_MANUAL):
        print(f"\n[{nombre_competencia}] Aviso: B4 de E01Pauta está vacío. Usando CAMPEON_REAL_MANUAL como fallback.")
    else:
        print(f"\n[{nombre_competencia}] Aviso: no hay campeón oficial en E01Pauta (B4 vacío). No se asignará Bono Campeón.")

    puntos_repartidos = calcular_puntos_repartidos(
        pautas_por_etapa=pautas_por_etapa,
        campeon_real_oficial=campeon_real_oficial,
        max_por_etapa=max_por_etapa,
        max_bonus=max_bonus,
    )
    porcentaje_avance = (puntos_repartidos / max_total * 100) if max_total else 0

    # 5) Ranking general (sin grupos)
    participantes = []

    for pid, info in datos.items():
        bono = 0
        if campeon_real_norm and norm(info.get("campeon_pred")) == campeon_real_norm:
            bono = BONUS_PTS

        total = sum(info["scores"].values()) + bono
        participantes.append((pid, info["nombre"], info["scores"], bono, total, info["errores"]))

    participantes.sort(key=lambda x: (-x[4], x[1].upper()))  # total desc, nombre asc

    print("\n" + "=" * ancho_tabla)
    print(f"Tabla de posiciones - {titulo_competencia}")
    print("=" * ancho_tabla)

    # Header fila 1
    header1 = l("Pos", W_POS) + " " + l("Nombre", W_NOMBRE) + " " + c("Total", W_TOTAL) + " "
    for e in etapas_ordenadas:
        header1 += c(ETIQUETAS_ETAPAS[e], W_ETAPA) + " "
    header1 += c(NOMBRE_COLUMNA_BONUS, W_BONUS)
    print(header1)

    # Header fila 2 (máximos)
    header2 = (" " * W_POS) + " " + (" " * W_NOMBRE) + " " + c(f"Max={max_total}", W_TOTAL) + " "
    for e in etapas_ordenadas:
        header2 += c(f"Max={max_por_etapa[e]}", W_ETAPA) + " "
    header2 += c(f"Max={max_bonus}", W_BONUS)
    print(header2)

    print("-" * len(header1))

    posiciones_participantes = calcular_posiciones_con_empate(participantes)

    # Filas
    for pos, (pid, nombre, scores, bono, total, errores) in zip(posiciones_participantes, participantes):
        row = l(pos, W_POS) + " " + l(nombre, W_NOMBRE) + " " + c(total, W_TOTAL) + " "
        for e in etapas_ordenadas:
            row += c(scores[e], W_ETAPA) + " "
        row += c(bono, W_BONUS)
        print(row)

        for err in errores:
            print(f"  -> ERROR: {err}")

    participantes_html = []
    for pos, (pid, nombre, scores, bono, total, errores) in zip(posiciones_participantes, participantes):
        participantes_html.append((pid, pos, nombre, scores, bono, total))

    final_oficial_finalizada = final_oficial_completa(pautas_por_etapa)
    mostrar_simulador_final = not final_oficial_finalizada
    simulacion_final_payload = (
        construir_simulacion_final_payload(datos, participantes_html)
        if mostrar_simulador_final
        else None
    )
    titulo_premios = (
        "Polla Familia"
        if nombre_competencia.lower() == "familia"
        else titulo_competencia
    )
    premios_payload = construir_premios_payload(
        titulo_competencia=titulo_premios,
        pozo_premios=pozo_premios,
        pautas_por_etapa=pautas_por_etapa,
        final_oficial_finalizada=final_oficial_finalizada,
    )

    participantes_select = [
        {"id": pid, "name": info["nombre"]}
        for pid, info in sorted(datos.items(), key=lambda x: x[1]["nombre"].upper())
    ]
    campeones_por_participante = {
        pid: valor_payload(info.get("campeon_pred"))
        for pid, info in datos.items()
    }
    etapas_comenzadas = {
        e: etapa_comenzada(pautas_por_etapa, e)
        for e in etapas_ordenadas
    }
    etapas_finalizadas = {
        e: etapa_finalizada(pautas_por_etapa, e)
        for e in etapas_ordenadas
    }
    podios_por_etapa = calcular_podios_por_etapa(
        participantes=participantes,
        etapas_ordenadas=etapas_ordenadas,
        etapas_finalizadas=etapas_finalizadas,
    )
    participantes_familiares_html = None
    podios_familiares = None

    if usar_ranking_familiar:
        try:
            datos_familias = cargar_mapa_familias(NOMBRES_PARTICIPANTES_PATH)
            for aviso in datos_familias["avisos"]:
                print(f"[{nombre_competencia}] Aviso familias: {aviso}")

            ranking_familiar_info = construir_ranking_familiar(
                participantes=participantes,
                datos_familias=datos_familias,
                etapas_ordenadas=etapas_ordenadas,
            )
            if ranking_familiar_info["cantidad_familias_excel"] != 7:
                print(
                    f"[{nombre_competencia}] Aviso familias: el Excel contiene "
                    f"{ranking_familiar_info['cantidad_familias_excel']} familias únicas; "
                    "se esperaban normalmente 7."
                )
            for nombre in ranking_familiar_info["participantes_sin_familia"]:
                print(
                    f"[{nombre_competencia}] Aviso familias: participante cargado sin "
                    f"familia asignada: {nombre}"
                )
            for nombre in ranking_familiar_info["nombres_excel_sin_coincidencia"]:
                print(
                    f"[{nombre_competencia}] Aviso familias: nombre del Excel sin "
                    f"participante cargado: {nombre}"
                )
            for familia in ranking_familiar_info["familias_sin_integrantes"]:
                print(
                    f"[{nombre_competencia}] Aviso familias: familia sin integrantes "
                    f"válidos: {familia}"
                )

            ranking_familiar = ranking_familiar_info["ranking"]
            posiciones_familiares = calcular_posiciones_con_empate(ranking_familiar)
            participantes_familiares_html = [
                (familia_id, pos, familia, scores, bono, total)
                for pos, (familia_id, familia, scores, bono, total, errores)
                in zip(posiciones_familiares, ranking_familiar)
            ]
            podios_familiares = calcular_podios_por_etapa(
                participantes=ranking_familiar,
                etapas_ordenadas=etapas_ordenadas,
                etapas_finalizadas=etapas_finalizadas,
                ranking_familiar=True,
            )

            print("\n" + "=" * ancho_tabla)
            print(f"Tabla de posiciones familiar - {titulo_competencia}")
            print("=" * ancho_tabla)
            print(
                "Pos | Familia | Total | Grupos | 16avos | Octavos | "
                "Cuartos | Semis | Final | Bono Campeón"
            )
            print("-" * ancho_tabla)
            for pos, (familia_id, familia, scores, bono, total, errores) in zip(
                posiciones_familiares, ranking_familiar
            ):
                puntajes = " | ".join(
                    formatear_puntaje_familiar(scores[e])
                    for e in etapas_ordenadas
                )
                print(
                    f"{pos} | {familia} | {formatear_puntaje_familiar(total)} | "
                    f"{puntajes} | {formatear_puntaje_familiar(bono)}"
                )
                integrantes = ranking_familiar_info["integrantes"].get(familia_id, [])
                print(f"{familia}: {', '.join(integrantes)}")
        except ValueError as e:
            print(
                f"[{nombre_competencia}] ERROR en nombres_participantes.xlsx: {e} "
                "Se generará solamente la tabla individual."
            )
        except Exception as e:
            print(
                f"[{nombre_competencia}] Aviso: no se pudo leer "
                f"nombres_participantes.xlsx ({e}). Se generará solamente la tabla individual."
            )
    etapas_select = [
        {
            "id": e,
            "label": etiqueta_etapa_larga(e),
            "show_bonus": ETAPAS[e]["tipo"] != "GRUPOS",
            "started": etapas_comenzadas[e],
        }
        for e in etapas_ordenadas
    ]
    detalles_ui = {}
    for pid, info in datos.items():
        detalles_ui[pid] = {}
        for etapa in etapas_ordenadas:
            if not etapas_comenzadas[etapa]:
                continue
            detalle = info.get("detalle_etapas", {}).get(etapa)
            if not detalle:
                continue
            detalles_ui[pid][etapa] = {
                "total": detalle["total_etapa"],
                "partidos": detalle["partidos"],
            }

    bonus_campeon_started = etapas_comenzadas.get("E01", False)
    bonus_campeon_participantes = []
    if bonus_campeon_started:
        for pid, info in sorted(datos.items(), key=lambda x: x[1]["nombre"].upper()):
            campeon_pred = info.get("campeon_pred")
            puntos_bono = (
                BONUS_PTS
                if campeon_real_norm and norm(campeon_pred) == campeon_real_norm
                else 0
            )
            bonus_campeon_participantes.append({
                "id": pid,
                "name": info["nombre"],
                "champion": valor_payload(campeon_pred),
                "points": puntos_bono,
            })

    payload_detalle = {
        "participants": participantes_select,
        "stages": etapas_select,
        "details": detalles_ui,
        "match_labels": {
            e: enfrentamientos_por_etapa.get(e, [])
            for e in etapas_ordenadas
        },
        "bonus_champion": {
            "started": bonus_campeon_started,
            "official_champion": valor_payload(campeon_real_oficial) if bonus_campeon_started else "",
            "participants": bonus_campeon_participantes,
        },
    }
    payload_resultados = construir_resultados_payload(
        etapas_ordenadas=etapas_ordenadas,
        pautas_por_etapa=pautas_por_etapa,
        enfrentamientos_detalle_por_etapa=enfrentamientos_detalle_por_etapa,
        calendario_por_etapa=calendario_por_etapa,
    )

    out_dir = os.path.join(OUTPUT_DIR, subcarpeta_salida)
    os.makedirs(out_dir, exist_ok=True)
    out_html = os.path.join(out_dir, "index.html")
    render_tabla_html(
        nombre_competencia=titulo_competencia,
        participantes=participantes_html,
        etapas_ordenadas=etapas_ordenadas,
        max_por_etapa=max_por_etapa,
        max_bonus=max_bonus,
        max_total=max_total,
        out_path=out_html,
        detalle_payload=payload_detalle,
        resultados_payload=payload_resultados,
        tendencias_payload=tendencias_payload,
        pronosticos_tabla_payload=pronosticos_tabla_payload,
        puntos_repartidos=puntos_repartidos,
        porcentaje_avance=porcentaje_avance,
        podios_por_etapa=podios_por_etapa,
        participantes_familiares=participantes_familiares_html,
        podios_familiares=podios_familiares,
        campeones_por_participante=campeones_por_participante,
        mostrar_simulador_final=mostrar_simulador_final,
        simulacion_final_payload=simulacion_final_payload,
        pozo_premios=pozo_premios,
        premios_payload=premios_payload,
    )
    print(f"[{nombre_competencia}] HTML generado: {out_html}")
    return True


def main():
    calendario_por_etapa = preparar_y_cargar_calendario()

    resultados = {}
    resultados["familia"] = generar_competencia(
        nombre_competencia="familia",
        nombre_carpeta_participantes="Participantes_Familia",
        subcarpeta_salida="familia",
        calendario_por_etapa=calendario_por_etapa,
        usar_ranking_familiar=True,
        pozo_premios=75000,
    )
    resultados["curso"] = generar_competencia(
        nombre_competencia="Segundos Medios",
        nombre_carpeta_participantes="Participantes_Curso",
        subcarpeta_salida="curso",
        calendario_por_etapa=calendario_por_etapa,
        usar_ranking_familiar=False,
        pozo_premios=84000,
    )

    out_portada = os.path.join(OUTPUT_DIR, "index.html")
    render_portada_html(out_portada)
    print(f"\nPortada generada: {out_portada}")

    if not any(resultados.values()):
        print("\nERROR: no se pudo generar ninguna competencia.")
        sys.exit(1)



if __name__ == "__main__":
    main()
